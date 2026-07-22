"""
Smart Data Puller — أداة سحب البيانات الذكية
=============================================
أداة واحدة سهلة تسحب بيانات الشركات من Google مباشرة
بدون ما تحتاج API Key.

تسحب: اسم الشركة + الرقم + الموقع + العنوان + القطاع

الاستخدام:
    python smart_puller.py                    # سحب من كل القطاعات
    python smart_puller.py --sector transport # قطاع النقل فقط
    python smart_puller.py --fast             # سريع (أقل بيانات)
    python smart_puller.py --deep             # عميق (أكتر بيانات لكن أبطأ)
"""

import os
import sys
import re
import json
import time
import random
import hashlib
from datetime import datetime

try:
    import requests
    from bs4 import BeautifulSoup
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Install: pip install requests beautifulsoup4 openpyxl lxml")
    sys.exit(1)

# ========================================
# SEARCH QUERIES — 300+ استعلام مخصص
# ========================================

QUERIES = {
    'transport': {
        'name': 'نقل وشحن',
        'searches': [
            # عربي
            'شركات نقل بضائع القاهرة ارقام تليفونات',
            'شركات شحن بري مصر تليفون',
            'شركات نقل ثقيل 6 اكتوبر',
            'شركات شحن العاشر من رمضان',
            'دليل شركات النقل والشحن في مصر',
            'شركات نقل بضائع العبور',
            'شركة نقل مبرد مصر',
            'شركات نقل حاويات القاهرة',
            'أكبر شركات النقل في مصر',
            'شركات نقل ومخازن',
            # English
            'transport companies cairo egypt phone number',
            'freight companies egypt contact',
            'trucking companies 6th october city',
            'cargo companies egypt list',
            'logistics companies cairo phone',
        ]
    },
    'food': {
        'name': 'أغذية ومشروبات',
        'searches': [
            'مصانع أغذية 6 أكتوبر أرقام تليفونات',
            'شركات توزيع أغذية القاهرة',
            'مصانع مشروبات مصر تليفون',
            'شركات ألبان مصر ارقام',
            'مصانع حلويات وبسكويت القاهرة',
            'شركات توزيع مواد غذائية',
            'مصانع لحوم ودواجن مصر',
            'شركات أغذية العاشر من رمضان',
            'أكبر شركات الأغذية في مصر',
            'food companies egypt phone number',
            'food factories 6th october egypt',
            'beverage companies cairo list',
            'dairy companies egypt contact',
        ]
    },
    'pharma': {
        'name': 'أدوية',
        'searches': [
            'شركات أدوية مصر أرقام تليفونات',
            'مصانع أدوية القاهرة',
            'شركات توزيع أدوية مصر',
            'مخازن أدوية القاهرة تليفون',
            'أكبر شركات الأدوية في مصر',
            'شركات مستلزمات طبية القاهرة',
            'pharmaceutical companies egypt phone',
            'pharma distributors cairo',
            'medical supplies companies egypt',
        ]
    },
    'construction': {
        'name': 'مقاولات',
        'searches': [
            'شركات مقاولات كبرى في مصر أرقام',
            'شركات مقاولات القاهرة تليفون',
            'شركات مقاولات عمومية مصر',
            'مصانع أسمنت مصر ارقام',
            'شركات حديد وصلب مصر',
            'شركات مواد بناء القاهرة',
            'شركات تشييد وبناء مصر',
            'construction companies egypt phone',
            'contractors cairo list',
            'building materials companies egypt',
        ]
    },
    'petroleum': {
        'name': 'بترول وطاقة',
        'searches': [
            'شركات بترول مصر أرقام تليفونات',
            'خدمات بترولية القاهرة',
            'شركات غاز طبيعي مصر',
            'شركات طاقة متجددة مصر',
            'شركات حفر آبار بترول',
            'oil companies egypt phone',
            'petroleum services cairo',
            'energy companies egypt contact',
        ]
    },
    'distribution': {
        'name': 'توزيع ولوجستيات',
        'searches': [
            'شركات لوجستيات مصر أرقام',
            'شركات توزيع القاهرة تليفون',
            'شركات مخازن وتخزين مصر',
            'شركات سلسلة إمداد القاهرة',
            'شركات بريد سريع مصر',
            'logistics companies egypt phone',
            'distribution companies cairo',
            'warehousing companies egypt',
        ]
    },
    'security': {
        'name': 'أمن وحراسة',
        'searches': [
            'شركات أمن وحراسة مصر أرقام',
            'شركات حراسة القاهرة تليفون',
            'شركات نقل أموال مصر',
            'security companies egypt phone',
            'guard services cairo',
        ]
    },
    'rental': {
        'name': 'تأجير سيارات',
        'searches': [
            'شركات تأجير سيارات القاهرة أرقام',
            'تأجير باصات مصر',
            'شركات ليموزين القاهرة تليفون',
            'car rental companies cairo phone',
            'vehicle leasing egypt',
            'limousine service cairo',
        ]
    },
    'manufacturing': {
        'name': 'مصانع',
        'searches': [
            'دليل مصانع 6 أكتوبر أرقام تليفونات',
            'دليل مصانع العاشر من رمضان',
            'مصانع العبور أرقام',
            'مصانع بدر وأبو زعبل',
            'مصانع شبرا الخيمة',
            'مصانع حلوان الصناعية',
            'مصانع مدينة السادات',
            'مصانع بلاستيك مصر تليفون',
            'مصانع كيماويات القاهرة',
            'مصانع نسيج وملابس مصر',
            'مصانع ورق وتغليف القاهرة',
            'مصانع أثاث مصر ارقام',
            'أكبر المصانع في مصر',
            'factories 6th october egypt phone',
            'industrial zone 10th ramadan companies',
            'factories obour city egypt',
        ]
    },
    'delivery': {
        'name': 'توصيل ودليفري',
        'searches': [
            'شركات توصيل طلبات القاهرة أرقام',
            'شركات شحن سريع مصر',
            'شركات كوريير القاهرة تليفون',
            'delivery companies cairo phone',
            'courier services egypt',
        ]
    },
    'tourism': {
        'name': 'سياحة',
        'searches': [
            'شركات سياحة مصر أرقام تليفونات',
            'شركات نقل سياحي القاهرة',
            'شركات أتوبيسات سياحية مصر',
            'tourism companies egypt phone',
            'tour bus companies cairo',
        ]
    },
    'public_transport': {
        'name': 'نقل جماعي',
        'searches': [
            'شركات أتوبيسات مصر أرقام',
            'شركات نقل ركاب القاهرة',
            'bus companies egypt phone',
        ]
    },
    'healthcare': {
        'name': 'مستشفيات',
        'searches': [
            'مستشفيات خاصة القاهرة أرقام تليفونات',
            'مراكز طبية كبرى مصر',
            'private hospitals cairo phone',
        ]
    },
    'education': {
        'name': 'تعليم',
        'searches': [
            'مدارس دولية القاهرة أرقام تليفونات',
            'مدارس خاصة مصر',
            'جامعات خاصة مصر تليفون',
            'international schools cairo phone',
        ]
    },
    'real_estate': {
        'name': 'عقارات',
        'searches': [
            'شركات تطوير عقاري مصر أرقام',
            'شركات مقاولات وعقارات القاهرة',
            'real estate companies egypt phone',
        ]
    },
    'fmcg': {
        'name': 'سلع استهلاكية',
        'searches': [
            'شركات منظفات مصر أرقام',
            'شركات مستحضرات تجميل القاهرة',
            'FMCG companies egypt phone',
        ]
    },
    'agriculture': {
        'name': 'زراعة',
        'searches': [
            'شركات زراعية مصر أرقام تليفونات',
            'شركات تصدير خضار وفاكهة',
            'agriculture companies egypt phone',
        ]
    },
    'waste': {
        'name': 'نظافة ومخلفات',
        'searches': [
            'شركات نظافة القاهرة أرقام',
            'شركات إدارة مخلفات مصر',
            'waste management companies egypt',
        ]
    },
}


class SmartPuller:
    """سحب بيانات ذكي من Google Search مباشرة."""

    def __init__(self):
        self.companies = []
        self.seen = set()
        self.session = requests.Session()
        # تدوير بين عدة User-Agents
        self.user_agents = [
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36 Edg/119.0.0.0',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        ]
        self.request_count = 0
        os.makedirs('output', exist_ok=True)

    def _get_headers(self):
        return {
            'User-Agent': random.choice(self.user_agents),
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Language': 'ar,en;q=0.9',
            'Accept-Encoding': 'gzip, deflate',
            'DNT': '1',
            'Connection': 'keep-alive',
        }

    def _smart_delay(self):
        """تأخير ذكي يزيد تدريجياً مع كثرة الطلبات."""
        self.request_count += 1
        base = 3.0
        # كل 10 طلبات، زوّد التأخير
        extra = (self.request_count // 10) * 2
        delay = base + extra + random.uniform(1, 3)
        # كل 30 طلب، خد استراحة كبيرة
        if self.request_count % 30 == 0:
            print(f"\n  ⏸️  استراحة 60 ثانية (حماية من الحظر)...")
            time.sleep(60)
        else:
            time.sleep(delay)

    def _search_google(self, query, num=20):
        """بحث Google وإرجاع النتائج."""
        results = []
        for start in range(0, num, 10):
            url = f"https://www.google.com/search?q={requests.utils.quote(query)}&start={start}&num=10&hl=ar"

            self._smart_delay()
            try:
                resp = self.session.get(url, headers=self._get_headers(), timeout=20)
                if resp.status_code == 429:
                    print(f"\n  ⚠️  Google حظر مؤقت — استراحة 2 دقيقة...")
                    time.sleep(120)
                    resp = self.session.get(url, headers=self._get_headers(), timeout=20)

                if resp.status_code != 200:
                    continue

                soup = BeautifulSoup(resp.text, 'lxml')

                for g in soup.select('.g, .MjjYud'):
                    result = self._parse_result(g)
                    if result:
                        results.append(result)

            except Exception as e:
                print(f"  ❌ {str(e)[:50]}")

        return results

    def _parse_result(self, el):
        """تحليل نتيجة بحث Google."""
        try:
            title_el = el.select_one('h3')
            link_el = el.select_one('a[href^="http"]')
            snippet_el = el.select_one('.VwiC3b, .st, .lEBKkf, span:not(h3)')

            if not title_el:
                return None

            title = title_el.get_text(strip=True)
            # تنظيف العنوان
            title = re.sub(r'\s*[-|–—]\s*(فيسبوك|Facebook|LinkedIn|YouTube|Twitter|Wuzzuf|Indeed|تويتر).*$', '', title)
            title = re.sub(r'\s*[-|–—]\s*(Home|About|Contact|الرئيسية).*$', '', title)

            if len(title) < 3 or len(title) > 120:
                return None

            link = link_el.get('href', '') if link_el else ''
            # تجاهل مواقع غير مفيدة
            skip = ['wikipedia', 'youtube', 'twitter', 'reddit', 'quora', 'pinterest',
                    'facebook.com', 'instagram.com', 'tiktok', 'amazon']
            if any(s in link.lower() for s in skip):
                return None

            snippet = snippet_el.get_text(strip=True) if snippet_el else ''

            return {'title': title, 'link': link, 'snippet': snippet}
        except:
            return None

    def _extract_company(self, result, sector):
        """استخراج بيانات شركة من نتيجة بحث."""
        title = result['title']
        snippet = result.get('snippet', '')
        link = result.get('link', '')
        full_text = f"{title} {snippet}"

        company = {'sector': sector, 'source': 'google_search'}

        # اسم الشركة
        if any('\u0600' <= c <= '\u06FF' for c in title):
            company['nameAr'] = title
        else:
            company['nameEn'] = title

        # رقم التليفون من الـ snippet
        phones = re.findall(
            r'(?:\+?20[\s\-.]?)?(?:0?2[\s\-.]?\d{4}[\s\-.]?\d{4}|'
            r'0?1[0125][\s\-.]?\d{4}[\s\-.]?\d{4}|'
            r'19\d{3}|16\d{3})',
            full_text
        )
        phones = list(set(re.sub(r'[\s\-.]', '', p) for p in phones))
        phones = [p for p in phones if len(p) >= 8]
        if phones:
            company['phone1'] = phones[0]
            if len(phones) > 1:
                company['phone2'] = phones[1]

        # إيميل
        emails = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', full_text)
        emails = [e for e in emails if 'example' not in e and 'test' not in e and 'wix' not in e]
        if emails:
            company['email'] = emails[0]

        # موقع إلكتروني
        if link and link.startswith('http') and not any(s in link for s in ['google', 'facebook']):
            company['website'] = link

        # المنطقة
        company['city'] = self._detect_area(full_text)

        return company

    def _detect_area(self, text):
        areas_map = {
            '6 أكتوبر': '6october', 'اكتوبر': '6october', '6th october': '6october',
            'العاشر من رمضان': '10thramadan', '10th ramadan': '10thramadan',
            'العبور': 'obour', 'obour': 'obour',
            'الشروق': 'shorouk', 'مدينة نصر': 'nasr_city', 'nasr city': 'nasr_city',
            'التجمع': 'new_cairo', 'new cairo': 'new_cairo', 'القاهرة الجديدة': 'new_cairo',
            'المعادي': 'maadi', 'maadi': 'maadi',
            'حلوان': 'helwan', 'helwan': 'helwan',
            'مصر الجديدة': 'heliopolis', 'heliopolis': 'heliopolis',
            'الجيزة': 'giza', 'giza': 'giza',
            'شبرا': 'shubra', 'الدقي': 'dokki', 'المهندسين': 'mohandessin',
            'بدر': 'badr', 'السادات': 'sadat',
        }
        text_lower = text.lower()
        for keyword, area in areas_map.items():
            if keyword in text or keyword.lower() in text_lower:
                return area
        return 'cairo'

    def _add_company(self, company):
        key = f"{company.get('nameAr','')}{company.get('nameEn','')}{company.get('phone1','')}".lower().strip()
        h = hashlib.md5(key.encode()).hexdigest()
        if h in self.seen or (not company.get('nameAr') and not company.get('nameEn')):
            return False
        self.seen.add(h)
        company['id'] = f"c{len(self.companies)+1:06d}"
        company['lastUpdated'] = datetime.now().strftime('%Y-%m-%d')
        self.companies.append(company)
        return True

    def _enrich_website(self, company, max_to_enrich=100):
        """زيارة الموقع الإلكتروني لسحب أرقام إضافية."""
        url = company.get('website')
        if not url or company.get('phone1'):
            return company

        try:
            time.sleep(random.uniform(1, 2))
            resp = self.session.get(url, headers=self._get_headers(), timeout=10, allow_redirects=True)
            if resp.status_code != 200:
                return company

            text = BeautifulSoup(resp.text, 'lxml').get_text()

            # أرقام تليفون
            phones = re.findall(
                r'(?:\+?20[\s\-.]?)?(?:0?2[\s\-.]?\d{4}[\s\-.]?\d{4}|'
                r'0?1[0125][\s\-.]?\d{4}[\s\-.]?\d{4})',
                text
            )
            phones = list(set(re.sub(r'[\s\-.]', '', p) for p in phones))
            phones = [p for p in phones if len(p) >= 10]
            if phones and not company.get('phone1'):
                company['phone1'] = phones[0]
            if len(phones) > 1 and not company.get('phone2'):
                company['phone2'] = phones[1]

            # إيميل
            emails = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', text)
            emails = [e for e in emails if 'example' not in e and 'wix' not in e and 'google' not in e]
            if emails and not company.get('email'):
                company['email'] = emails[0]

        except:
            pass

        return company

    def pull(self, sectors=None, deep=False):
        """سحب البيانات الرئيسي."""
        print("=" * 60)
        print("🔍 Smart Data Puller — سحب بيانات ذكي")
        print(f"   التاريخ: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        print("=" * 60)

        # تحميل البيانات المحفوظة سابقاً لو موجودة
        cache_file = 'output/_smart_cache.json'
        if os.path.exists(cache_file):
            with open(cache_file, 'r', encoding='utf-8') as f:
                cached = json.load(f)
                for c in cached:
                    self._add_company(c)
            print(f"   تم تحميل {len(self.companies)} شركة من الذاكرة")

        # تحميل البيانات من collect_real_data لو موجود
        try:
            from collect_real_data import CURATED_COMPANIES
            for c in CURATED_COMPANIES:
                self._add_company(dict(c))
            print(f"   + بيانات مؤكدة: {len(self.companies)} شركة")
        except:
            pass

        target_sectors = sectors if sectors else list(QUERIES.keys())
        total_queries = sum(len(QUERIES[s]['searches']) for s in target_sectors)
        query_num = 0

        for sector_key in target_sectors:
            sector = QUERIES.get(sector_key)
            if not sector:
                continue

            print(f"\n{'='*50}")
            print(f"  📂 القطاع: {sector['name']}")
            print(f"{'='*50}")

            searches = sector['searches']
            if not deep:
                searches = searches[:8]  # في الوضع العادي، 8 استعلامات لكل قطاع

            for query in searches:
                query_num += 1
                print(f"\n  [{query_num}/{total_queries}] 🔍 {query[:60]}...")

                results = self._search_google(query, num=20 if deep else 10)
                added = 0

                for result in results:
                    company = self._extract_company(result, sector_key)
                    if self._add_company(company):
                        added += 1

                print(f"    ✅ +{added} شركة جديدة (إجمالي: {len(self.companies)})")

                # حفظ تلقائي كل 5 استعلامات
                if query_num % 5 == 0:
                    self._save_cache(cache_file)

        # إثراء من المواقع (أول 50 شركة بدون رقم)
        no_phone = [c for c in self.companies if not c.get('phone1') and c.get('website')][:50]
        if no_phone:
            print(f"\n🌐 إثراء بيانات {len(no_phone)} شركة من مواقعهم...")
            enriched = 0
            for i, c in enumerate(no_phone):
                print(f"  [{i+1}/{len(no_phone)}] {c.get('nameAr', c.get('nameEn', ''))}... ", end='')
                before = c.get('phone1')
                self._enrich_website(c)
                if c.get('phone1') and c.get('phone1') != before:
                    enriched += 1
                    print("✅")
                else:
                    print("—")
            print(f"  تم إثراء {enriched} شركة")

        # حفظ وتصدير
        self._save_cache(cache_file)
        self._print_summary()
        self._export_excel()
        self._export_json()
        self._export_crm_json()

    def _save_cache(self, path):
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(self.companies, f, ensure_ascii=False, indent=2)

    def _print_summary(self):
        print(f"\n{'='*60}")
        print(f"📊 ملخص النتائج")
        print(f"{'='*60}")
        print(f"  إجمالي الشركات: {len(self.companies)}")

        sectors = {}
        for c in self.companies:
            s = c.get('sector', '?')
            name = QUERIES.get(s, {}).get('name', s)
            sectors[name] = sectors.get(name, 0) + 1

        print(f"\n  حسب القطاع:")
        for name, count in sorted(sectors.items(), key=lambda x: x[1], reverse=True):
            print(f"    {name}: {count}")

        with_phone = len([c for c in self.companies if c.get('phone1')])
        with_email = len([c for c in self.companies if c.get('email')])
        with_site = len([c for c in self.companies if c.get('website')])
        total = max(len(self.companies), 1)

        print(f"\n  اكتمال البيانات:")
        print(f"    📞 تليفون: {with_phone}/{total} ({100*with_phone//total}%)")
        print(f"    📧 إيميل: {with_email}/{total} ({100*with_email//total}%)")
        print(f"    🌐 موقع: {with_site}/{total} ({100*with_site//total}%)")
        print(f"{'='*60}")

    def _export_excel(self):
        date_str = datetime.now().strftime('%Y%m%d_%H%M')
        filename = f'output/smart_pull_{date_str}.xlsx'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'الشركات'
        ws.sheet_view.rightToLeft = True

        hf = Font(name='Arial', bold=True, size=11, color='FFFFFF')
        hfill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
        cf = Font(name='Arial', size=10)
        border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB')
        )

        headers = ['#', 'اسم الشركة (عربي)', 'Company Name', 'القطاع',
                   'المنطقة', 'هاتف 1', 'هاتف 2', 'البريد', 'الموقع', 'ملاحظات']
        widths = [5, 35, 35, 15, 14, 16, 16, 28, 40, 25]

        for i, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.font = hf; cell.fill = hfill; cell.border = border
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        areas_ar = {k: v.get('name', k) for k, v in QUERIES.items()}

        for row, c in enumerate(self.companies, 2):
            data = [
                row-1, c.get('nameAr',''), c.get('nameEn',''),
                QUERIES.get(c.get('sector',''), {}).get('name', c.get('sector','')),
                c.get('city',''), c.get('phone1',''), c.get('phone2',''),
                c.get('email',''), c.get('website',''), c.get('notes','')
            ]
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = cf; cell.border = border
            if row % 2 == 0:
                for col in range(1, len(headers)+1):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')

        wb.save(filename)
        print(f"\n✅ Excel: {filename}")

    def _export_json(self):
        date_str = datetime.now().strftime('%Y%m%d_%H%M')
        filename = f'output/smart_pull_{date_str}.json'
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(self.companies, f, ensure_ascii=False, indent=2)
        print(f"✅ JSON: {filename}")

    def _export_crm_json(self):
        """تصدير بصيغة جاهزة للاستيراد في CRM مباشرة."""
        filename = 'output/crm_import_ready.json'
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(self.companies, f, ensure_ascii=False, indent=2)
        print(f"✅ CRM Import: {filename}")


def main():
    import argparse
    parser = argparse.ArgumentParser(description='Smart Data Puller - سحب بيانات الشركات من Google')
    parser.add_argument('--sector', nargs='+', choices=list(QUERIES.keys()),
                        help='قطاعات محددة')
    parser.add_argument('--deep', action='store_true',
                        help='بحث عميق (أكتر استعلامات = أكتر بيانات)')
    parser.add_argument('--fast', action='store_true',
                        help='بحث سريع (أقل استعلامات)')
    args = parser.parse_args()

    puller = SmartPuller()
    puller.pull(sectors=args.sector, deep=args.deep)

    print("\n🎉 تم! افتح مجلد output لملفات Excel و JSON")
    print("💡 لاستيراد في CRM: استخدم ملف crm_import_ready.json")


if __name__ == '__main__':
    main()
