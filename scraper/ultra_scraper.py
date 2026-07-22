# -*- coding: utf-8 -*-
"""
Ultra Scraper v3 — Exhaustive Fleet-Focused Company Census
===========================================================
Advanced Google Maps scraper with:
  - undetected-chromedriver (anti-detection)
  - Geographic Grid Segmentation (bypass 120-result limit)
  - Browser Auto-Restart (session rotation)
  - Exponential Backoff with Jitter
  - Zero-Results Detection (block detection)
  - Atomic JSON writes (no corruption)
  - Config-driven (scraper_config.json)

Usage:
    python ultra_scraper.py                    # Full run (exhaustive)
    python ultra_scraper.py --resume           # Continue from last run
    python ultra_scraper.py --test 5           # Test mode (5 searches only)
    python ultra_scraper.py --no-grid          # Disable grid, use city names
    python ultra_scraper.py --visible          # Show browser (not headless)
"""

import os, sys, re, json, time, random, hashlib, argparse, tempfile, shutil
from datetime import datetime
from contextlib import contextmanager

sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

try:
    import undetected_chromedriver as uc
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    USE_UC = True
except ImportError:
    try:
        from selenium import webdriver
        from selenium.webdriver.common.by import By
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        USE_UC = False
        print("⚠️ undetected-chromedriver not found, using standard Selenium (higher block risk)")
    except ImportError:
        print("❌ Run: pip install undetected-chromedriver"); sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

SCRAPER_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(SCRAPER_DIR, 'output')
CACHE_FILE = os.path.join(OUTPUT_DIR, '_ultra_cache.json')
PROGRESS_FILE = os.path.join(OUTPUT_DIR, '_ultra_progress.json')
CONFIG_FILE = os.path.join(OUTPUT_DIR, 'scraper_config.json')
LOG_FILE = os.path.join(OUTPUT_DIR, 'scraper.log')
LOCK_FILE = os.path.join(OUTPUT_DIR, '_crm_write.lock')

# ============================================================
# UTILITY: Atomic JSON Save (prevents corruption)
# ============================================================

@contextmanager
def file_lock(lock_path=LOCK_FILE, timeout=120, poll=0.15):
    """Cross-process lock for CRM/cache/progress read-merge-write cycles."""
    os.makedirs(os.path.dirname(lock_path), exist_ok=True)
    lock_file = open(lock_path, 'a+b')
    acquired = False
    start = time.time()
    try:
        if os.name == 'nt':
            import msvcrt
            while True:
                try:
                    lock_file.seek(0)
                    msvcrt.locking(lock_file.fileno(), msvcrt.LK_NBLCK, 1)
                    acquired = True
                    break
                except OSError:
                    if time.time() - start >= timeout:
                        raise TimeoutError(f"Timed out waiting for output lock: {lock_path}")
                    time.sleep(poll)
        else:
            import fcntl
            while True:
                try:
                    fcntl.flock(lock_file.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
                    acquired = True
                    break
                except OSError:
                    if time.time() - start >= timeout:
                        raise TimeoutError(f"Timed out waiting for output lock: {lock_path}")
                    time.sleep(poll)
        yield
    finally:
        try:
            if acquired:
                if os.name == 'nt':
                    import msvcrt
                    lock_file.seek(0)
                    msvcrt.locking(lock_file.fileno(), msvcrt.LK_UNLCK, 1)
                else:
                    import fcntl
                    fcntl.flock(lock_file.fileno(), fcntl.LOCK_UN)
        finally:
            lock_file.close()

def atomic_json_save(filepath, data):
    """Write JSON atomically with retries. Call inside file_lock for shared CRM files."""
    dir_name = os.path.dirname(filepath) or '.'
    os.makedirs(dir_name, exist_ok=True)
    fd, tmp_path = tempfile.mkstemp(dir=dir_name, suffix='.tmp', prefix='.save_')
    try:
        with os.fdopen(fd, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False)
        last_error = None
        for attempt in range(10):
            try:
                os.replace(tmp_path, filepath)
                return
            except PermissionError as e:
                last_error = e
                time.sleep(0.2 + attempt * 0.15)
        raise last_error or PermissionError(f"Could not replace {filepath}")
    finally:
        try:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
        except Exception:
            pass



def log(msg):
    """Log to console and file."""
    ts = datetime.now().strftime('%H:%M:%S')
    line = f"[{ts}] {msg}"
    print(line)
    try:
        with open(LOG_FILE, 'a', encoding='utf-8') as f:
            f.write(line + '\n')
    except:
        pass


# ============================================================
# CONFIG LOADER
# ============================================================
def load_config():
    """Load scraper configuration from JSON file."""
    if not os.path.exists(CONFIG_FILE):
        print(f"⚠️ Config file not found: {CONFIG_FILE}. Using defaults.")
        return None
    with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)


# ============================================================
# GEOGRAPHIC GRID GENERATOR
# ============================================================
def generate_grid(bounds, rows, cols):
    """
    Generate a grid of center coordinates for geographic segmentation.
    Each cell = a search center point with lat,lng.
    """
    lat_step = (bounds['north'] - bounds['south']) / rows
    lng_step = (bounds['east'] - bounds['west']) / cols
    cells = []
    for r in range(rows):
        for c in range(cols):
            center_lat = bounds['south'] + lat_step * (r + 0.5)
            center_lng = bounds['west'] + lng_step * (c + 0.5)
            cells.append({
                'lat': round(center_lat, 5),
                'lng': round(center_lng, 5),
                'label': f"grid_{r}_{c}"
            })
    return cells


# ============================================================
# MAIN SCRAPER CLASS
# ============================================================
class UltraScraper:
    def __init__(self, config=None, test_mode=0, no_grid=False, visible=False):
        self.config = config or {}
        self.test_mode = test_mode
        self.no_grid = no_grid
        self.visible = visible
        
        # Performance settings
        perf = self.config.get('performance', {})
        self.delay_min = perf.get('delay_min', 4)
        self.delay_max = perf.get('delay_max', 9)
        self.restart_every = perf.get('restart_browser_every', 50)
        self.max_zero_restart = perf.get('max_zero_results_before_restart', 3)
        self.max_zero_backoff = perf.get('max_zero_results_before_backoff', 5)
        self.backoff_base = perf.get('backoff_base_seconds', 120)
        self.backoff_max = perf.get('backoff_max_seconds', 600)
        self.scroll_rounds = perf.get('scroll_rounds', 10)
        self.small_result_scroll_rounds = perf.get('small_result_scroll_rounds', 2)
        self.render_wait_max = perf.get('render_wait_max_seconds', 6)
        self.post_render_sleep_min = perf.get('post_render_sleep_min', 0.25)
        self.post_render_sleep_max = perf.get('post_render_sleep_max', 0.65)
        self.scroll_wait_min = perf.get('scroll_wait_min', 0.25)
        self.scroll_wait_max = perf.get('scroll_wait_max', 0.65)
        self.page_load_timeout = perf.get('page_load_timeout_seconds', 18)
        self.headless = perf.get('headless', True) and not visible
        
        # State
        self.companies = []
        self.seen_hashes = set()
        self.completed_searches = set()
        self.stats = {}
        self.search_count = 0
        self.consecutive_zeros = 0
        self.backoff_level = 0
        self.driver = None
        
        os.makedirs(OUTPUT_DIR, exist_ok=True)

    # --------------------------------------------------------
    # BROWSER MANAGEMENT
    # --------------------------------------------------------
    def _create_driver(self):
        """Create a new browser instance with anti-detection."""
        user_agents = [
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:128.0) Gecko/20100101 Firefox/128.0",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.5 Safari/605.1.15",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/127.0.0.0 Safari/537.36",
        ]
        ua = random.choice(user_agents)
        
        if USE_UC:
            def get_uc_options():
                opt = uc.ChromeOptions()
                opt.add_argument('--lang=ar')
                opt.add_argument(f'--window-size={random.choice([1920,1366,1440])},{random.choice([1080,768,900])}')
                if self.headless:
                    opt.add_argument('--headless=new')
                return opt
            
            try:
                driver = uc.Chrome(options=get_uc_options(), use_subprocess=True)
                try:
                    driver.set_page_load_timeout(self.page_load_timeout)
                except Exception:
                    pass
                return driver
            except Exception as e:
                err_msg = str(e)
                # Try to parse the actual browser version from error message
                match = re.search(r'Current browser version is (\d+)', err_msg)
                if match:
                    major = int(match.group(1))
                    log(f"  ⚠️ Version mismatch. Forcing undetected-chromedriver to Chrome version {major}")
                    try:
                        driver = uc.Chrome(options=get_uc_options(), use_subprocess=True, version_main=major)
                        try:
                            driver.set_page_load_timeout(self.page_load_timeout)
                        except Exception:
                            pass
                        return driver
                    except Exception as e2:
                        log(f"  ❌ Failed with version_main override: {e2}")
                
                log("  ⚠️ undetected-chromedriver failed completely, falling back to standard Selenium")
                return self._create_standard_selenium(ua)
        else:
            return self._create_standard_selenium(ua)

    def _create_standard_selenium(self, ua):
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options as StdOptions
        options = StdOptions()
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--lang=ar')
        options.add_argument(f'user-agent={ua}')
        options.add_argument(f'--window-size={random.choice([1920,1366,1440])},{random.choice([1080,768,900])}')
        options.add_argument('--disable-blink-features=AutomationControlled')
        options.add_experimental_option('excludeSwitches', ['enable-automation'])
        options.add_experimental_option('useAutomationExtension', False)
        if self.headless:
            options.add_argument('--headless=new')
        driver = webdriver.Chrome(options=options)
        try:
            driver.set_page_load_timeout(self.page_load_timeout)
        except Exception:
            pass
        driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        return driver

    def _restart_browser(self, reason="scheduled"):
        """Close and reopen browser with new fingerprint."""
        log(f"  🔄 Restarting browser ({reason})...")
        if self.driver:
            try:
                self.driver.quit()
            except:
                pass
            self.driver = None
        
        # Random wait before restart
        wait = random.uniform(10, 25)
        log(f"  ⏳ Waiting {wait:.0f}s before new session...")
        time.sleep(wait)
        
        self.driver = self._create_driver()
        self.search_count = 0
        self.consecutive_zeros = 0
        self.backoff_level = 0
        log("  ✅ New browser session started")

    # --------------------------------------------------------
    # DEDUPLICATION
    # --------------------------------------------------------
    def _hash(self, company):
        """Create a dedup hash for a company. Prioritizes the unique Google Maps place key to preserve multiple branch locations."""
        place_key = company.get('google_maps_place_key') or self._extract_maps_place_key(company.get('google_maps_url'))
        if place_key:
            return hashlib.md5(place_key.lower().strip().encode()).hexdigest()
            
        # Fallback to name + phone + address if place key is missing
        key = f"{company.get('nameAr','')}{company.get('nameEn','')}{company.get('phone1','')}{company.get('address','')}".lower().strip()
        return hashlib.md5(key.encode()).hexdigest()

    def _clean_text_field(self, value):
        """Remove map UI artifacts while preserving Arabic/English business text."""
        if not value:
            return value
        text = str(value)
        text = ''.join(ch for ch in text if not (0xE000 <= ord(ch) <= 0xF8FF))
        text = re.sub(r'^[\s·•\-\u202d\u202c]+', '', text)
        text = re.sub(r'[\u202a-\u202e]', '', text)
        text = re.sub(r'\s+', ' ', text).strip()
        if text in {'·', '-', '—'}:
            return ''
        return text

    def _normalize_phone(self, value):
        """Normalize common Egyptian phone formats without inventing missing digits."""
        if not value:
            return ''
        digits = re.sub(r'\D', '', str(value))
        if digits.startswith('0020'):
            digits = digits[2:]
        if digits.startswith('20') and len(digits) in (11, 12):
            digits = '0' + digits[2:]
        if len(digits) == 10 and digits.startswith('1'):
            digits = '0' + digits
        if len(digits) == 9 and digits.startswith('2'):
            digits = '0' + digits
        return digits

    def _extract_maps_place_key(self, url):
        if not url:
            return ''
        match = re.search(r'!1s([^!/?&#]+)', url)
        if match:
            return match.group(1)
        match = re.search(r'!16s%2Fg%2F([^!/?&#]+)', url)
        if match:
            return match.group(1)
        return ''

    def _extract_maps_coordinates(self, url):
        if not url:
            return None, None
        match = re.search(r'!3d(-?\d+(?:\.\d+)?)!4d(-?\d+(?:\.\d+)?)', url)
        if not match:
            return None, None
        try:
            return float(match.group(1)), float(match.group(2))
        except ValueError:
            return None, None

    def _target_result_bounds(self):
        grid_cfg = self.config.get('grid', {})
        return grid_cfg.get('result_bounds') or grid_cfg.get('cairo_giza_result_bounds') or {
            'south': 29.55, 'north': 30.45,
            'west': 30.70, 'east': 31.80
        }

    def _is_outside_target_area(self, company):
        if self.config.get('mode') != 'tire_lead_generation_cairo_giza':
            return False

        lat = company.get('latitude')
        lng = company.get('longitude')
        if lat is None or lng is None:
            lat, lng = self._extract_maps_coordinates(company.get('google_maps_url'))

        if lat is not None and lng is not None:
            bounds = self._target_result_bounds()
            margin = float(self.config.get('grid', {}).get('result_bounds_margin', 0.02))
            if (
                lat < float(bounds.get('south', -90)) - margin or
                lat > float(bounds.get('north', 90)) + margin or
                lng < float(bounds.get('west', -180)) - margin or
                lng > float(bounds.get('east', 180)) + margin
            ):
                return True
            return False

        text = ' '.join([
            str(company.get('nameAr') or ''),
            str(company.get('nameEn') or ''),
            str(company.get('sector_details') or ''),
            str(company.get('address') or ''),
        ]).lower()
        outside_keywords = [
            'المحلة', 'طلخا', 'المنصورة', 'الدقهلية', 'الغربية', 'طنطا',
            'كفر الشيخ', 'دمياط', 'بورسعيد',
            'الشرقية', 'الزقازيق', 'بلبيس', 'الصالحية',
            'المنوفية', 'السادات', 'بنها', 'القليوبية', 'الخانكة',
            'الفيوم', 'بني سويف', 'المنيا', 'أسيوط', 'اسيوط',
            'dakahlia', 'gharbia', 'mansoura', 'tanta', 'mahalla',
            'sharqia', 'zagazig', 'portsaid',
        ]
        return any(keyword in text for keyword in outside_keywords)

    def _lead_text(self, company):
        return ' '.join([
            str(company.get('nameAr') or ''),
            str(company.get('nameEn') or ''),
            str(company.get('sector_details') or ''),
            str(company.get('address') or ''),
        ]).lower()

    def _has_tire_buyer_signal(self, company):
        """Require a real public signal that the place operates vehicles or fleet-heavy work."""
        text = self._lead_text(company)
        buyer_keywords = [
            'شركة نقل', 'شركة النقل', 'خدمة نقل', 'خدمات نقل', 'وسائل النقل',
            'نقل وتخزين', 'نقل العمال', 'نقل ثقيل', 'النقل الثقيل',
            'إدارة النقل', 'ادارة النقل', 'قطاع النقل', 'نقليات', 'للنقل',
            'نقل البضائع', 'أساطيل', 'اساطيل',
            'شحن', 'لوجست', 'توصيل', 'توزيع', 'موزع', 'اسطول', 'أسطول', 'شاحن', 'تريلا',
            'اتوبيس', 'أتوبيس', 'حافلات', 'سفريات', 'معدات ثقيلة', 'تاجير معدات',
            'تأجير معدات', 'تأجير سيارات', 'مقطورات', 'صهاريج', 'حاويات',
            'مقاولات', 'خرسانة', 'مصنع', 'مواد غذائية', 'بترول', 'محاجر',
            'ليموزين', 'تاكسي', 'رحلات', 'سياحة', 'سياحى', 'سياحية', 'سيارات', 'ملاكي', 'ملاكى', 'اجرة', 'أجرة',
            'transport', 'trucking', 'truck', 'fleet', 'shipping', 'logistics',
            'freight', 'cargo', 'delivery', 'distribution', 'bus', 'coach',
            'heavy equipment', 'construction', 'trailer', 'container',
            'limousine', 'taxi', 'travel', 'tourism', 'rent', 'rental', 'distributor', 'car', 'vehicle', 'passenger'
        ]
        return any(keyword in text for keyword in buyer_keywords)

    def _tire_lead_profile(self, company):
        """Score how likely this company is to repeatedly buy tires."""
        sector_key = company.get('sector', '')
        sector_cfg = self.config.get('sectors', {}).get(sector_key, {})
        priority = sector_cfg.get('priority', company.get('priority', 'C'))
        base_by_priority = {'A+': 72, 'A': 62, 'B': 50, 'C': 38}
        score = base_by_priority.get(priority, 40)
        reasons = []

        text = self._lead_text(company)

        strong_keywords = {
            'نقل ثقيل': 15, 'النقل الثقيل': 15, 'إدارة النقل': 12,
            'ادارة النقل': 12, 'قطاع النقل': 11, 'نقل وتخزين': 10,
            'نقل العمال': 10, 'وسائل النقل': 8,
            'اسطول': 14, 'أسطول': 14, 'معدات ثقيلة': 13,
            'تريلا': 12, 'شاحن': 11, 'شركة نقل': 10, 'شركة النقل': 10,
            'خدمة نقل': 9, 'خدمات نقل': 9,
            'نقل البضائع': 10, 'نقليات': 9, 'مقطورات': 9, 'صهاريج': 10,
            'حاويات': 9, 'لوجست': 8, 'شحن': 7, 'توزيع': 7, 'توصيل': 6,
            'اتوبيس': 9, 'أتوبيس': 9, 'حافلات': 9, 'مقاولات': 7,
            'خرسانة': 8, 'مصنع': 5, 'truck': 11, 'fleet': 13,
            'trucking': 12, 'transport': 8, 'freight': 8, 'cargo': 7,
            'logistics': 8, 'delivery': 6, 'distribution': 7, 'bus': 9,
            'heavy equipment': 12, 'construction': 7, 'trailer': 9,
        }
        for kw, weight in strong_keywords.items():
            if kw in text:
                score += weight
                reasons.append(kw)

        if self._has_tire_buyer_signal(company):
            score += 6
            reasons.append('buyer-signal')
        else:
            score -= 45
            reasons.append('no-buyer-signal')
        if company.get('phone1'):
            score += 4
            reasons.append('phone')
        if company.get('google_maps_url'):
            score += 3
            reasons.append('maps')
        if company.get('rating'):
            score += 2
            reasons.append('rating')
        non_buyer_keywords = {
            'صيانة': 28, 'تصليح': 28, 'خراطة': 32, 'حدادة': 28, 'ميكانيكا': 32,
            'ميكانيكي': 32, 'سوست': 30, 'بنشر': 35, 'قطع غيار': 40,
            'مركز صيانة': 35, 'مركز خدمة سيارات': 35, 'مركز تصليح': 35,
            'car service': 35, 'repair': 30, 'mechanic': 35, 'workshop': 35,
            'spare parts': 40, 'auto parts': 40,
        }
        for kw, penalty in non_buyer_keywords.items():
            if kw in text:
                score -= penalty
                reasons.append(f'non-buyer:{kw}')

        if self._is_outside_target_area(company):
            score -= 35
            reasons.append('outside-target-area')
        if company.get('operating_status') == 'permanently_closed':
            score -= 45
            reasons.append('closed')

        score = max(0, min(100, score))
        if score >= 85:
            lead_priority = 'A+'
        elif score >= 72:
            lead_priority = 'A'
        elif score >= 58:
            lead_priority = 'B'
        else:
            lead_priority = 'C'

        return score, lead_priority, ', '.join(dict.fromkeys(reasons))

    def _prepare_company(self, company):
        """Clean, normalize, and attach tire-sales lead fields."""
        for field in ('nameAr', 'nameEn', 'sector_details', 'address', 'working_hours'):
            if field in company:
                cleaned = self._clean_text_field(company.get(field))
                if cleaned:
                    company[field] = cleaned
                else:
                    company.pop(field, None)

        for field in ('phone1', 'phone2', 'mobile'):
            if field in company:
                normalized = self._normalize_phone(company.get(field))
                if normalized:
                    company[field] = normalized
                else:
                    company.pop(field, None)

        phone_sources = company.get('phone_sources') if isinstance(company.get('phone_sources'), dict) else {}
        for field in ('phone1', 'phone2', 'mobile'):
            phone = company.get(field)
            if phone:
                phone_sources.setdefault(phone, {
                    'source': 'google_maps_card',
                    'confidence': 0.92,
                    'verified': True
                })
        if phone_sources:
            company['phone_sources'] = phone_sources
            company['phone_verified'] = True
            company['phone_confidence'] = max(
                float(v.get('confidence', 0)) for v in phone_sources.values()
                if isinstance(v, dict)
            )

        maps_key = self._extract_maps_place_key(company.get('google_maps_url'))
        if maps_key:
            company['google_maps_place_key'] = maps_key
        lat, lng = self._extract_maps_coordinates(company.get('google_maps_url'))
        if lat is not None and lng is not None:
            company['latitude'] = lat
            company['longitude'] = lng

        sector_cfg = self.config.get('sectors', {}).get(company.get('sector', ''), {})
        if sector_cfg.get('priority'):
            company['sector_priority'] = sector_cfg.get('priority')
        if self.config.get('mode') == 'tire_lead_generation_cairo_giza':
            score, priority, reasons = self._tire_lead_profile(company)
            company['tire_need_score'] = score
            company['tire_priority'] = priority
            company['tire_lead_reason'] = reasons
        return company

    def _is_bad_tire_lead(self, company):
        """Reject non-buyers: tire sellers, workshops, individual drivers, and out-of-area results."""
        if self.config.get('mode') != 'tire_lead_generation_cairo_giza':
            return False
        if self._is_outside_target_area(company):
            return True
        text = ' '.join([
            str(company.get('nameAr') or ''),
            str(company.get('nameEn') or ''),
            str(company.get('sector_details') or ''),
            str(company.get('address') or ''),
        ]).lower()
        bad_keywords = [
            'كاوتش', 'إطارات', 'اطارات', 'فرد كاوتش', 'جنط', 'جنوط',
            'قطع غيار', 'اكسسوارات سيارات', 'بطاريات', 'زيوت سيارات',
            'ورشة', 'بنشر', 'مركز خدمة سيارات', 'مركز صيانة سيارات',
            'مركز صيانة', 'مركز تصليح', 'تصليح سيارات', 'صيانة سيارات',
            'ميكانيكا', 'ميكانيكي', 'خراطة', 'حدادة', 'سوست', 'فرامل',
            'كهرباء سيارات', 'خدمة الديزل', 'سائق نقل', 'سائق ',
            'مكتب بريد', 'فرع بنك', 'atm',
            'tires', 'tyres', 'tire shop', 'tyre shop', 'auto parts',
            'spare parts', 'car service', 'mechanic', 'oil change',
            'workshop', 'repair shop', 'driver', 'post office',
        ]
        if any(keyword in text for keyword in bad_keywords):
            return True
        return not self._has_tire_buyer_signal(company)

    def _add(self, company):
        """Add a company if not duplicate. Returns True if added."""
        company = self._prepare_company(company)
        if self._is_bad_tire_lead(company):
            return False
        if not company.get('nameAr') and not company.get('nameEn'):
            return False
        h = self._hash(company)
        if h in self.seen_hashes:
            return False
        self.seen_hashes.add(h)
        place_key = company.get('google_maps_place_key') or self._extract_maps_place_key(company.get('google_maps_url'))
        if place_key:
            h_id = hashlib.md5(place_key.lower().strip().encode()).hexdigest()[:12]
            company['id'] = f"u_{h_id}"
        else:
            key_str = f"{company.get('nameAr','')}{company.get('nameEn','')}{company.get('phone1','')}{company.get('address','')}".lower().strip()
            h_id = hashlib.md5(key_str.encode()).hexdigest()[:12]
            company['id'] = f"u_{h_id}"
        company['lastUpdated'] = datetime.now().strftime('%Y-%m-%d')
        self.companies.append(company)
        sector = company.get('sector', 'other')
        self.stats[sector] = self.stats.get(sector, 0) + 1
        return True

    # --------------------------------------------------------
    # SEARCH GOOGLE MAPS
    # --------------------------------------------------------
    def _search_maps(self, query, location_label, sector_key, lat=None, lng=None):
        """
        Search Google Maps for a query.
        If lat/lng provided, search at that coordinate (grid mode).
        Otherwise, search by city name (legacy mode).
        """
        if lat is not None and lng is not None:
            search_id = f"{query}|{lat},{lng}"
            url = f"https://www.google.com/maps/search/{query.replace(' ', '+')}/@{lat},{lng},14z"
        else:
            search_id = f"{query}|{location_label}"
            full_query = f"{query} {location_label}"
            url = f"https://www.google.com/maps/search/{full_query.replace(' ', '+')}"

        if search_id in self.completed_searches:
            return 0

        try:
            self.driver.get(url)
            
            # Smart wait loop: bounded by config for faster grid scans
            start_wait = time.time()
            is_empty = False
            is_blocked = False
            
            no_results_patterns = [
                'لم يتم العثور', 'no results found', 'cannot find', 'could not find',
                'verify the spelling', 'تأكد من كتابة'
            ]
            
            while time.time() - start_wait < self.render_wait_max:
                page_source = self.driver.page_source or ''
                
                # Check block
                if 'unusual traffic' in page_source.lower() or 'captcha' in page_source.lower():
                    is_blocked = True
                    break
                    
                # Fetch visible text for empty cell detection to avoid matching JS boilerplate
                try:
                    visible_text = self.driver.find_element(By.TAG_NAME, 'body').text or ''
                except:
                    visible_text = ''
                    
                # Check empty cell
                if any(p in visible_text.lower() for p in no_results_patterns):
                    is_empty = True
                    break
                    
                # Check if results are rendered
                cards = self.driver.find_elements(By.CSS_SELECTOR, 'div.Nv2PK')
                if cards:
                    break
                    
                time.sleep(0.5)
                
            if is_blocked:
                log("  🚨 CAPTCHA/Block detected!")
                return -1
                
            if is_empty:
                self.completed_searches.add(search_id)
                return 0

            # Brief human-like randomized sleep after rendering
            time.sleep(random.uniform(self.post_render_sleep_min, self.post_render_sleep_max))

            # Scroll results panel to load more (with adaptive scroll bounds)
            try:
                panel = self.driver.find_element(By.CSS_SELECTOR, 'div[role="feed"]')
                last_height = self.driver.execute_script('return arguments[0].scrollHeight', panel)
                
                # Check initial number of results
                initial_cards = self.driver.find_elements(By.CSS_SELECTOR, 'div.Nv2PK')
                
                # Early duplicate exit check: if 100% of first page is duplicate, skip
                if initial_cards:
                    all_duplicates = True
                    for card in initial_cards:
                        try:
                            text = card.text or ''
                            lines = [l.strip() for l in text.split('\n') if l.strip()]
                            if lines:
                                name = lines[0]
                                temp_comp = {'phone1': ''}
                                if any('\u0600' <= c <= '\u06FF' for c in name):
                                    temp_comp['nameAr'] = name
                                else:
                                    temp_comp['nameEn'] = name
                                
                                h = self._hash(temp_comp)
                                if h not in self.seen_hashes:
                                    all_duplicates = False
                                    break
                        except:
                            all_duplicates = False
                            break
                    if all_duplicates and len(initial_cards) >= 3:
                        log(f"  ⏩ Early exit: all {len(initial_cards)} visible results are duplicates. Skipping scroll.")
                        self.completed_searches.add(search_id)
                        return 0
                
                # Dynamic scroll limiting: if very few results initially, don't waste time scrolling 10 times
                max_scrolls = self.scroll_rounds
                if len(initial_cards) < 4:
                    max_scrolls = min(self.small_result_scroll_rounds, max_scrolls)
                
                for scroll_i in range(max_scrolls):
                    self.driver.execute_script(
                        'arguments[0].scrollTop = arguments[0].scrollHeight', panel
                    )
                    time.sleep(random.uniform(self.scroll_wait_min, self.scroll_wait_max))
                    
                    # Exit early if height hasn't changed (end of listings)
                    new_height = self.driver.execute_script('return arguments[0].scrollHeight', panel)
                    if new_height == last_height:
                        break
                    last_height = new_height
                    
                    # Check if "end of list" marker appeared
                    try:
                        end_marker = self.driver.find_element(By.CSS_SELECTOR, 'span.HlvSq')
                        if end_marker:
                            break
                    except:
                        pass
            except:
                pass

            # Extract business cards
            added = 0
            cards = self.driver.find_elements(By.CSS_SELECTOR, 'div.Nv2PK')

            for card in cards:
                try:
                    text = card.text or ''
                    if not text or len(text) < 5:
                        continue

                    lines = [l.strip() for l in text.split('\n') if l.strip()]
                    if not lines:
                        continue

                    # Strict ad-listing check (verify first 3 lines of card text)
                    is_ad = False
                    for line in lines[:3]:
                        if line in ('إعلان', 'Ad', 'Sponsored', 'ممول', 'مُموَّل', 'إعلانات', 'Ads'):
                            is_ad = True
                            break
                    if is_ad:
                        continue

                    # Prioritize extracting the business name from the aria-label of the main card link
                    name = None
                    try:
                        link_el = card.find_element(By.CSS_SELECTOR, 'a.hfpxzc')
                        name = link_el.get_attribute('aria-label')
                    except:
                        pass
                    
                    if not name:
                        name = lines[0]
                    else:
                        name = name.strip()

                    if not name or len(name) < 3 or len(name) > 120:
                        continue

                    # Skip UI action words
                    skip_words = ['اتجاهات', 'الحصول على', 'الانتقال إلى',
                                  'موقع ويب', 'اتصال هاتف', 'حفظ', 'مشاركة', 'نجمة']
                    if any(w in name.lower() for w in skip_words) or name.lower() in ('مميز', 'featured'):
                        continue

                    company = {
                        'sector': sector_key,
                        'source': 'google_maps',
                        'city': location_label,
                    }

                    # Name (Arabic or English)
                    if any('\u0600' <= c <= '\u06FF' for c in name):
                        company['nameAr'] = name
                    else:
                        company['nameEn'] = name

                    # Rating
                    for line in lines[1:4]:
                        match = re.search(r'(\d\.\d)\s*\(([^)]+)\)', line)
                        if match:
                            company['rating'] = match.group(1)
                            company['reviews_count'] = match.group(2).strip()
                            break
                        elif re.match(r'^\d\.\d$', line):
                            company['rating'] = line
                            break

                    # Operating status
                    company['operating_status'] = 'active'
                    if any(k in text for k in ['مغلق نهائياً', 'Permanently closed']):
                        company['operating_status'] = 'permanently_closed'
                    elif any(k in text for k in ['مغلق مؤقتاً', 'Temporarily closed']):
                        company['operating_status'] = 'temporarily_closed'

                    # Working hours
                    status_kw = ['مفتوح الآن', 'يغلق في', 'يفتح في', 'مفتوح ٢٤ ساعة',
                                 'مغلق', 'Open now', 'Closed', 'Opens', 'Open 24 hours']
                    for line in lines:
                        if any(kw in line for kw in status_kw) and ' · ' not in line \
                                and not re.search(r'\d\.\d', line):
                            company['working_hours'] = line.strip()
                            break

                    # Category & Address
                    for line in lines:
                        if ' · ' in line and not any(k in line for k in
                                ['يفتح', 'يغلق', 'مفتوح', 'مغلق', 'Open', 'Closed']):
                            parts = line.split(' · ')
                            company['sector_details'] = parts[0].strip()
                            if len(parts) > 1:
                                company['address'] = ' · '.join(p.strip() for p in parts[1:])
                            break

                    # Phone numbers
                    phone_re = r'(?:\+?20[\s\-.]?)?(?:0?2[\s\-.]?\d{3,4}[\s\-.]?\d{4}|0?1[0125][\s\-.]?\d{3,4}[\s\-.]?\d{4}|19\d{3}|16\d{3})'
                    phones = re.findall(phone_re, text)
                    if phones:
                        clean_phones = []
                        for p in phones:
                            cp = re.sub(r'[\s\-.]', '', p)
                            if cp not in clean_phones:
                                clean_phones.append(cp)
                        if len(clean_phones) >= 1: company['phone1'] = clean_phones[0]
                        if len(clean_phones) >= 2: company['phone2'] = clean_phones[1]
                        if len(clean_phones) >= 3: company['mobile'] = clean_phones[2]

                    # Google Maps URL
                    try:
                        link_el = card.find_element(By.CSS_SELECTOR, 'a.hfpxzc')
                        href = link_el.get_attribute('href')
                        if href:
                            company['google_maps_url'] = href
                    except:
                        pass

                    if self._add(company):
                        added += 1

                except Exception:
                    continue

            self.completed_searches.add(search_id)
            return added

        except Exception as e:
            log(f"  ⚠️ Search error: {e}")
            err_msg = str(e).lower()
            if any(k in err_msg for k in ['invalid session id', 'invalid session', 'session deleted', 'no such window', 'chrome not reachable', 'disconnected']):
                return -1
            return 0

    # --------------------------------------------------------
    # SAFE BASELINE TEST (detect block)
    # --------------------------------------------------------
    def _is_blocked(self):
        """Check if blocked by looking at the page source for CAPTCHA indicators."""
        try:
            page_source = self.driver.page_source or ''
            if 'unusual traffic' in page_source.lower() or 'captcha' in page_source.lower():
                return True
            return False
        except:
            return True

    # --------------------------------------------------------
    # PERSISTENCE
    # --------------------------------------------------------
    def save_progress(self, status='running'):
        """Save companies/progress under one cross-process lock, preserving enriched fields."""
        with file_lock():
            merged_companies = self.companies
            if os.path.exists(CACHE_FILE):
                try:
                    with open(CACHE_FILE, 'r', encoding='utf-8') as f:
                        disk_companies = json.load(f)

                    merged_map = {}

                    def get_key(c):
                        return (
                            c.get('google_maps_place_key')
                            or c.get('google_maps_url')
                            or c.get('id')
                            or hashlib.md5(
                                f"{c.get('nameAr','')}{c.get('nameEn','')}{c.get('phone1','')}".lower().strip().encode()
                            ).hexdigest()
                        )

                    for c in disk_companies:
                        merged_map[get_key(c)] = c

                    for c in self.companies:
                        key = get_key(c)
                        if key in merged_map:
                            existing = merged_map[key]
                            for k, v in c.items():
                                if v is None or v == '':
                                    continue
                                if k == 'phone_sources' and isinstance(v, dict) and isinstance(existing.get(k), dict):
                                    combined = dict(v)
                                    combined.update(existing[k])
                                    existing[k] = combined
                                else:
                                    existing[k] = v
                        else:
                            merged_map[key] = c

                    merged_companies = list(merged_map.values())
                    try:
                        merged_companies.sort(key=lambda x: x.get('id', ''))
                    except Exception:
                        pass
                    self.companies = merged_companies
                except Exception as e:
                    log(f"  ⚠️ Error merging during save_progress: {e}")

            atomic_json_save(CACHE_FILE, merged_companies)
            crm_file = os.path.join(OUTPUT_DIR, 'crm_import_ready.json')
            atomic_json_save(crm_file, merged_companies)

            progress = {
                'total': len(merged_companies),
                'with_phone': sum(1 for c in merged_companies if c.get('phone1')),
                'target': 5000,
                'completed_searches': list(self.completed_searches),
                'hashes': list(self.seen_hashes),
                'stats': self.stats,
                'status': status,
                'timestamp': datetime.now().isoformat()
            }
            atomic_json_save(PROGRESS_FILE, progress)

    def load_progress(self):
        """Load previous progress under the same lock used by writers."""
        with file_lock(timeout=30):
            if not os.path.exists(CACHE_FILE):
                return False
            try:
                with open(CACHE_FILE, 'r', encoding='utf-8') as f:
                    self.companies = json.load(f)
                if os.path.exists(PROGRESS_FILE):
                    with open(PROGRESS_FILE, 'r', encoding='utf-8') as f:
                        progress = json.load(f)
                        self.completed_searches = set(progress.get('completed_searches', []))
                        self.seen_hashes = set(progress.get('hashes', []))
                        self.stats = progress.get('stats', {})
                if not self.seen_hashes:
                    for c in self.companies:
                        self.seen_hashes.add(self._hash(c))
                log(f"  📂 Resumed: {len(self.companies):,} companies, {len(self.completed_searches):,} searches done")
                return True
            except Exception as e:
                log(f"  ⚠️ Error loading progress: {e}")
                return False

    # --------------------------------------------------------
    # MAIN RUN LOOP    # --------------------------------------------------------
    # MAIN RUN LOOP
    # --------------------------------------------------------
    def run(self, resume=False):
        """Main scraping loop."""
        config = self.config
        
        # Build sector list from config
        sectors = {}
        config_sectors = config.get('sectors', {})
        for key, s in config_sectors.items():
            if s.get('enabled', True):
                sectors[key] = s
        
        # Build location list from config
        locations = []
        config_areas = config.get('focus_areas', {})
        for area_key, area in config_areas.items():
            if area.get('enabled', True):
                for city in area.get('cities', []):
                    if city not in locations:
                        locations.append(city)
        
        # Build grid if enabled
        grid_cfg = config.get('grid', {})
        use_grid = grid_cfg.get('enabled', True) and not self.no_grid
        grid_cells = []
        if use_grid:
            bounds = grid_cfg.get('cairo_giza_bounds', {
                'south': 29.85, 'north': 30.20,
                'west': 31.05, 'east': 31.50
            })
            rows = grid_cfg.get('rows', 7)
            cols = grid_cfg.get('cols', 9)
            grid_cells = generate_grid(bounds, rows, cols)
        
        # Calculate total searches
        total_queries = sum(len(s.get('queries', [])) for s in sectors.values())
        if use_grid:
            total_searches = total_queries * len(grid_cells)
        else:
            total_searches = total_queries * len(locations)
        
        print("=" * 65)
        print(f"🚀 ULTRA SCRAPER v3 — Exhaustive Fleet Census")
        print(f"   Sectors: {len(sectors)} | Queries: {total_queries}")
        if use_grid:
            print(f"   Grid: {len(grid_cells)} cells ({grid_cfg.get('rows',7)}×{grid_cfg.get('cols',9)})")
        else:
            print(f"   Cities: {len(locations)}")
        print(f"   Total searches: {total_searches:,}")
        print(f"   Anti-detection: {'undetected-chromedriver' if USE_UC else 'standard Selenium'}")
        print(f"   Browser restart every: {self.restart_every} searches")
        if self.test_mode:
            print(f"   ⚡ TEST MODE: {self.test_mode} searches only")
        print("=" * 65)

        if resume:
            self.load_progress()
        
        # Start browser
        self.driver = self._create_driver()
        log("  ✅ Browser started\n")
        
        search_num = 0
        test_limit = self.test_mode if self.test_mode > 0 else float('inf')
        
        # Sort sectors by priority (A+ first)
        priority_order = {'A+': 0, 'A': 1, 'B': 2, 'C': 3}
        sorted_sectors = sorted(
            sectors.items(),
            key=lambda x: priority_order.get(x[1].get('priority', 'C'), 3)
        )
        
        try:
            for sector_key, sector in sorted_sectors:
                if search_num >= test_limit:
                    break
                    
                priority = sector.get('priority', 'C')
                sector_name = sector.get('name', sector_key)
                
                print(f"\n{'='*55}")
                print(f"  🏭 [{priority}] {sector_name}")
                print(f"{'='*55}")
                
                for query in sector.get('queries', []):
                    if search_num >= test_limit:
                        break
                    
                    # Choose iteration: grid cells or city names
                    if use_grid:
                        search_targets = [
                            (cell['label'], cell['lat'], cell['lng'])
                            for cell in grid_cells
                        ]
                    else:
                        search_targets = [
                            (city, None, None)
                            for city in locations
                        ]
                    
                    for location_label, lat, lng in search_targets:
                        if search_num >= test_limit:
                            break
                        
                        # Check search_id to skip completed
                        if lat is not None:
                            sid = f"{query}|{lat},{lng}"
                        else:
                            sid = f"{query}|{location_label}"
                        if sid in self.completed_searches:
                            continue
                        
                        search_num += 1
                        self.search_count += 1
                        
                        # Display progress
                        loc_display = location_label if not lat else f"({lat:.3f},{lng:.3f})"
                        print(f"  [{search_num}] {query} — {loc_display}... ", end='', flush=True)
                        
                        # Execute search
                        added = self._search_maps(query, location_label, sector_key, lat, lng)
                        
                        if added == -1:
                            # Block detected
                            print("🚨 BLOCKED!")
                            self._restart_browser("block detected")
                            continue
                        
                        print(f"+{added} (total: {len(self.companies):,})")
                        
                        # Track consecutive zeros
                        if added == 0:
                            self.consecutive_zeros += 1
                        else:
                            self.consecutive_zeros = 0
                            self.backoff_level = 0
                        
                        # Zero-result detection: restart browser
                        if self.consecutive_zeros >= self.max_zero_restart and \
                                self.consecutive_zeros < self.max_zero_backoff:
                            if not self._is_blocked():
                                # Genuine zero results, not blocked
                                pass
                            else:
                                self._restart_browser(
                                    f"{self.consecutive_zeros} consecutive zeros — possibly blocked"
                                )
                        
                        # Exponential backoff on sustained zeros (only if actually blocked)
                        if self.consecutive_zeros >= self.max_zero_backoff:
                            if not self._is_blocked():
                                # Genuine zero results (e.g. desert or empty coordinate cell), not a block
                                # Reset counter to continue scanning next coordinates directly
                                self.consecutive_zeros = 0
                                self.backoff_level = 0
                            else:
                                self.backoff_level += 1
                                backoff_time = min(
                                    self.backoff_base * (2 ** (self.backoff_level - 1)),
                                    self.backoff_max
                                )
                                # Add jitter
                                backoff_time += random.uniform(0, 30)
                                log(f"  ⏸️ Backoff level {self.backoff_level}: waiting {backoff_time:.0f}s...")
                                time.sleep(backoff_time)
                                self._restart_browser("backoff recovery")
                                self.consecutive_zeros = 0
                        
                        # Scheduled browser restart
                        if self.search_count >= self.restart_every:
                            self.save_progress()
                            self._restart_browser("scheduled rotation")
                        
                        # Auto-save every 5 searches
                        if search_num % 5 == 0:
                            self.save_progress()
                            print(f"  💾 Synced: {len(self.companies):,} companies")
                
                # Save after each sector
                self.save_progress()
                log(f"  ✅ Sector '{sector_name}' done — {len(self.companies):,} total")
        
        except KeyboardInterrupt:
            print("\n\n  ⚠️ Stopped by user — saving...")
        
        finally:
            if self.driver:
                try:
                    self.driver.quit()
                except:
                    pass
        self.save_progress(status='stopped')
        
        # Run Google Maps details enricher automatically to pull missing phone numbers and details
        log("  ⚡ Running Google Maps Details Enricher to pull missing phone numbers and details...")
        try:
            import subprocess
            enrich_cmd = [sys.executable, '-X', 'utf8', 'maps_detail_enricher.py', '--limit', '2500']
            subprocess.run(enrich_cmd, creationflags=0x08000000)
            # Reload companies after enrichment to ensure Excel and summary are up-to-date
            self.load_progress()
        except Exception as e:
            log(f"  ⚠️ Details Enricher failed: {e}")
            
        self.export_excel()
        self.print_summary()

    # --------------------------------------------------------
    # EXPORT
    # --------------------------------------------------------
    def export_excel(self):
        """Export to Excel with formatted headers."""
        if not openpyxl:
            return
        
        date_str = datetime.now().strftime('%Y%m%d_%H%M')
        xlsx_file = os.path.join(OUTPUT_DIR, f'ULTRA_{date_str}.xlsx')
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'الشركات'
        ws.sheet_view.rightToLeft = True

        hf = Font(name='Arial', bold=True, size=11, color='FFFFFF')
        hfill = PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid')
        cf = Font(name='Arial', size=10)
        border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )

        headers = ['#', 'اسم الشركة (عربي)', 'Company Name', 'القطاع',
                    'المدينة', 'هاتف 1', 'هاتف 2', 'العنوان', 'التقييم', 'المصدر']
        widths = [7, 45, 35, 18, 15, 16, 16, 45, 8, 12]

        for i, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.font = hf; cell.fill = hfill; cell.border = border
            ws.column_dimensions[get_column_letter(i)].width = w

        for idx, c in enumerate(self.companies, 1):
            row = idx + 1
            vals = [
                idx,
                c.get('nameAr', ''),
                c.get('nameEn', ''),
                c.get('sector', ''),
                c.get('city', ''),
                c.get('phone1', ''),
                c.get('phone2', ''),
                c.get('address', ''),
                c.get('rating', ''),
                c.get('source', ''),
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = cf
                cell.border = border

        ws.auto_filter.ref = ws.dimensions
        wb.save(xlsx_file)
        log(f"  📊 Excel: {xlsx_file}")

    def print_summary(self):
        """Print final summary."""
        print(f"\n{'='*65}")
        print(f"  🏁 SCRAPING COMPLETE")
        print(f"{'='*65}")
        print(f"  Total companies: {len(self.companies):,}")
        print(f"  With phone: {sum(1 for c in self.companies if c.get('phone1')):,}")
        print(f"  Searches completed: {len(self.completed_searches):,}")
        print(f"\n  📊 By sector:")
        for sector, count in sorted(self.stats.items(), key=lambda x: -x[1]):
            print(f"    {sector}: {count}")
        print(f"{'='*65}")


# ============================================================
# MAIN
# ============================================================
def main():
    parser = argparse.ArgumentParser(description='Ultra Scraper v3')
    parser.add_argument('--resume', action='store_true', help='Resume from last run')
    parser.add_argument('--test', type=int, default=0, help='Test mode: run N searches only')
    parser.add_argument('--no-grid', action='store_true', help='Disable geographic grid')
    parser.add_argument('--visible', action='store_true', help='Show browser (not headless)')
    args = parser.parse_args()

    config = load_config()
    if not config:
        print("❌ Cannot run without config. Create scraper_config.json first.")
        sys.exit(1)

    scraper = UltraScraper(
        config=config,
        test_mode=args.test,
        no_grid=args.no_grid,
        visible=args.visible
    )
    scraper.run(resume=args.resume or True)  # Always resume by default


if __name__ == '__main__':
    main()
