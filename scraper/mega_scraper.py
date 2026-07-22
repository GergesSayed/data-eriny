"""
Egypt MEGA Scraper — Multi-Source Company Collector
====================================================
Scrapes 5,000-30,000 Egyptian companies from ALL available public sources:

Sources:
  1. Yellow Pages Egypt (yellowpages.com.eg) — Thousands of listings
  2. Dalil Egypt (dalil.com) — Egyptian business directory
  3. Google Search — Company discovery via targeted queries
  4. Egyptian Exchange (EGX) — ~250 listed companies with full data
  5. Wuzzuf Jobs — Companies hiring drivers/fleet managers = have fleets
  6. Kompass-style directories — Industrial company listings
  7. Facebook/Instagram Business Pages — via Google search
  8. Industrial Zone directories — 6th October, 10th Ramadan, etc.

Usage:
    python mega_scraper.py                          # Run all sources
    python mega_scraper.py --source yellowpages     # Yellow Pages only
    python mega_scraper.py --source google          # Google search only
    python mega_scraper.py --source wuzzuf          # Wuzzuf jobs only
    python mega_scraper.py --source egx             # EGX listed companies
    python mega_scraper.py --source directories     # Business directories
    python mega_scraper.py --max-companies 5000     # Stop at 5000
    python mega_scraper.py --resume                 # Resume from last run
"""

import os
import sys
import re
import json
import csv
import time
import random
import hashlib
import argparse
import logging
from datetime import datetime
from urllib.parse import quote_plus, urljoin, urlparse
from concurrent.futures import ThreadPoolExecutor, as_completed

try:
    import requests
    from bs4 import BeautifulSoup
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"Missing: {e}\nInstall: pip install requests beautifulsoup4 openpyxl lxml")
    sys.exit(1)

# ============================================================
# CONFIGURATION
# ============================================================

OUTPUT_DIR = 'output'
PROGRESS_FILE = os.path.join(OUTPUT_DIR, '_progress.json')
COMPANIES_CACHE = os.path.join(OUTPUT_DIR, '_companies_cache.json')

# Rate limiting
MIN_DELAY = 1.5
MAX_DELAY = 3.5
MAX_RETRIES = 3

# All sectors with Arabic/English names and search keywords
SECTORS = {
    'transport': {
        'ar': 'نقل وشحن', 'en': 'Transport & Shipping',
        'yp_cats': ['transport', 'shipping', 'freight', 'cargo', 'trucking'],
        'google_queries_ar': ['شركة نقل', 'شركة شحن', 'نقل بضائع', 'نقل ثقيل', 'شحن بري'],
        'google_queries_en': ['transport company egypt', 'freight company cairo', 'trucking company egypt', 'cargo egypt']
    },
    'food': {
        'ar': 'أغذية ومشروبات', 'en': 'Food & Beverages',
        'yp_cats': ['food', 'beverages', 'dairy', 'bakery', 'meat', 'restaurants-suppliers'],
        'google_queries_ar': ['مصنع أغذية', 'شركة مشروبات', 'توزيع أغذية', 'مصنع حلويات', 'شركة ألبان'],
        'google_queries_en': ['food factory egypt', 'beverage company egypt', 'food manufacturer cairo']
    },
    'pharma': {
        'ar': 'أدوية', 'en': 'Pharmaceuticals',
        'yp_cats': ['pharmaceutical', 'medicines', 'medical-supplies'],
        'google_queries_ar': ['شركة أدوية', 'مصنع أدوية', 'توزيع أدوية', 'مستلزمات طبية'],
        'google_queries_en': ['pharmaceutical company egypt', 'pharma factory cairo', 'medical supplies egypt']
    },
    'construction': {
        'ar': 'مقاولات', 'en': 'Construction',
        'yp_cats': ['construction', 'contractors', 'building-materials', 'cement', 'steel'],
        'google_queries_ar': ['شركة مقاولات', 'مقاولات عمومية', 'مواد بناء', 'شركة تشييد'],
        'google_queries_en': ['construction company egypt', 'contractor cairo', 'building company egypt']
    },
    'petroleum': {
        'ar': 'بترول وطاقة', 'en': 'Oil & Energy',
        'yp_cats': ['petroleum', 'oil', 'gas', 'energy'],
        'google_queries_ar': ['شركة بترول', 'خدمات بترولية', 'شركة طاقة', 'غاز طبيعي'],
        'google_queries_en': ['oil company egypt', 'petroleum services egypt', 'energy company cairo']
    },
    'distribution': {
        'ar': 'توزيع ولوجستيات', 'en': 'Distribution & Logistics',
        'yp_cats': ['logistics', 'distribution', 'warehousing', 'supply-chain'],
        'google_queries_ar': ['شركة توزيع', 'لوجستيات', 'مخازن', 'سلسلة إمداد'],
        'google_queries_en': ['logistics company egypt', 'distribution company cairo', 'warehouse egypt']
    },
    'security': {
        'ar': 'أمن وحراسة', 'en': 'Security Services',
        'yp_cats': ['security', 'guard-services'],
        'google_queries_ar': ['شركة أمن', 'حراسة', 'خدمات أمنية', 'نقل أموال'],
        'google_queries_en': ['security company egypt', 'guard services cairo']
    },
    'rental': {
        'ar': 'تأجير سيارات', 'en': 'Car Rental',
        'yp_cats': ['car-rental', 'limousine', 'vehicle-leasing'],
        'google_queries_ar': ['تأجير سيارات', 'إيجار سيارات', 'ليموزين', 'تأجير باصات'],
        'google_queries_en': ['car rental egypt', 'vehicle leasing cairo', 'limousine service egypt']
    },
    'manufacturing': {
        'ar': 'مصانع', 'en': 'Manufacturing',
        'yp_cats': ['factories', 'industrial', 'manufacturing', 'plastics', 'chemicals', 'textiles', 'paper'],
        'google_queries_ar': ['مصنع', 'مصانع العاشر', 'مصانع 6 أكتوبر', 'مصانع العبور', 'صناعات'],
        'google_queries_en': ['factory egypt', 'manufacturer cairo', 'industrial company egypt']
    },
    'delivery': {
        'ar': 'توصيل ودليفري', 'en': 'Delivery & Courier',
        'yp_cats': ['courier', 'delivery', 'express'],
        'google_queries_ar': ['شركة توصيل', 'دليفري', 'شحن سريع', 'بريد سريع'],
        'google_queries_en': ['delivery company egypt', 'courier cairo', 'express shipping egypt']
    },
    'tourism': {
        'ar': 'سياحة ونقل سياحي', 'en': 'Tourism & Travel',
        'yp_cats': ['tourism', 'travel', 'tour-operators'],
        'google_queries_ar': ['شركة سياحة', 'نقل سياحي', 'أتوبيس سياحي', 'سفر وسياحة'],
        'google_queries_en': ['tourism company egypt', 'travel agency cairo', 'tour bus egypt']
    },
    'public_transport': {
        'ar': 'نقل جماعي', 'en': 'Public Transport',
        'yp_cats': ['bus', 'public-transport'],
        'google_queries_ar': ['شركة أتوبيس', 'نقل جماعي', 'نقل ركاب', 'باصات'],
        'google_queries_en': ['bus company egypt', 'public transport cairo']
    },
    'healthcare': {
        'ar': 'مستشفيات ورعاية صحية', 'en': 'Healthcare',
        'yp_cats': ['hospitals', 'clinics', 'medical-centers'],
        'google_queries_ar': ['مستشفى خاص', 'مركز طبي', 'مستشفيات القاهرة'],
        'google_queries_en': ['private hospital egypt', 'medical center cairo']
    },
    'education': {
        'ar': 'تعليم', 'en': 'Education',
        'yp_cats': ['schools', 'universities', 'education'],
        'google_queries_ar': ['مدرسة خاصة', 'مدرسة دولية', 'جامعة خاصة', 'معهد'],
        'google_queries_en': ['international school egypt', 'private school cairo', 'university egypt']
    },
    'fmcg': {
        'ar': 'سلع استهلاكية', 'en': 'FMCG',
        'yp_cats': ['consumer-goods', 'cosmetics', 'detergents', 'personal-care'],
        'google_queries_ar': ['شركة منظفات', 'مستحضرات تجميل', 'سلع استهلاكية'],
        'google_queries_en': ['FMCG company egypt', 'consumer goods cairo']
    },
    'agriculture': {
        'ar': 'زراعة وتصدير', 'en': 'Agriculture & Export',
        'yp_cats': ['agriculture', 'farming', 'agricultural-equipment'],
        'google_queries_ar': ['شركة زراعية', 'تصدير زراعي', 'شركة تصدير خضار وفاكهة'],
        'google_queries_en': ['agriculture company egypt', 'farm export egypt']
    },
    'telecom': {
        'ar': 'اتصالات', 'en': 'Telecommunications',
        'yp_cats': ['telecommunications', 'telecom'],
        'google_queries_ar': ['شركة اتصالات', 'خدمات اتصالات'],
        'google_queries_en': ['telecom company egypt']
    },
    'real_estate': {
        'ar': 'عقارات وتطوير', 'en': 'Real Estate',
        'yp_cats': ['real-estate', 'property-developers'],
        'google_queries_ar': ['شركة عقارات', 'تطوير عقاري', 'شركة تشييد وتطوير'],
        'google_queries_en': ['real estate developer egypt', 'property company cairo']
    },
    'waste_management': {
        'ar': 'نظافة وإدارة مخلفات', 'en': 'Waste Management',
        'yp_cats': ['waste-management', 'recycling', 'cleaning'],
        'google_queries_ar': ['شركة نظافة', 'إدارة مخلفات', 'تدوير', 'شركة جمع قمامة'],
        'google_queries_en': ['waste management egypt', 'cleaning company cairo']
    },
}

# Greater Cairo areas with Google Maps search variations
AREAS = {
    'cairo': {'ar': 'القاهرة', 'en': 'Cairo'},
    'giza': {'ar': 'الجيزة', 'en': 'Giza'},
    'nasr_city': {'ar': 'مدينة نصر', 'en': 'Nasr City'},
    'heliopolis': {'ar': 'مصر الجديدة', 'en': 'Heliopolis'},
    'maadi': {'ar': 'المعادي', 'en': 'Maadi'},
    'new_cairo': {'ar': 'التجمع الخامس', 'en': 'New Cairo'},
    '6october': {'ar': '6 أكتوبر', 'en': '6th October'},
    '10thramadan': {'ar': 'العاشر من رمضان', 'en': '10th of Ramadan'},
    'obour': {'ar': 'العبور', 'en': 'Obour'},
    'shorouk': {'ar': 'الشروق', 'en': 'Shorouk'},
    'helwan': {'ar': 'حلوان', 'en': 'Helwan'},
    'shubra': {'ar': 'شبرا الخيمة', 'en': 'Shubra El Kheima'},
    'badr': {'ar': 'بدر', 'en': 'Badr City'},
    'sadat': {'ar': 'السادات', 'en': 'Sadat City'},
    'dokki': {'ar': 'الدقي', 'en': 'Dokki'},
    'mohandessin': {'ar': 'المهندسين', 'en': 'Mohandessin'},
    'zamalek': {'ar': 'الزمالك', 'en': 'Zamalek'},
    'downtown': {'ar': 'وسط البلد', 'en': 'Downtown Cairo'},
    'ain_shams': {'ar': 'عين شمس', 'en': 'Ain Shams'},
    'abbassia': {'ar': 'العباسية', 'en': 'Abbassia'},
}

# ============================================================
# LOGGING SETUP
# ============================================================

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler(os.path.join(OUTPUT_DIR, 'scraper.log'), encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger('MegaScraper')


# ============================================================
# MEGA SCRAPER CLASS
# ============================================================

class MegaScraper:
    """Multi-source scraper for Egyptian companies."""

    def __init__(self, max_companies=30000):
        self.max_companies = max_companies
        self.companies = []
        self.seen_hashes = set()
        self.stats = {source: 0 for source in
                      ['curated', 'yellowpages', 'google', 'egx', 'wuzzuf',
                       'directories', 'facebook', 'industrial_zones']}
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
                          '(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept-Language': 'ar,en;q=0.9',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        })
        os.makedirs(OUTPUT_DIR, exist_ok=True)

    def _delay(self):
        time.sleep(random.uniform(MIN_DELAY, MAX_DELAY))

    def _request(self, url, **kwargs):
        for attempt in range(MAX_RETRIES):
            try:
                self._delay()
                resp = self.session.get(url, timeout=20, **kwargs)
                if resp.status_code == 200:
                    return resp
                elif resp.status_code == 429:
                    logger.warning(f"Rate limited, waiting 30s...")
                    time.sleep(30)
                elif resp.status_code == 403:
                    logger.warning(f"Blocked (403) on {urlparse(url).netloc}")
                    return None
                else:
                    logger.debug(f"HTTP {resp.status_code}: {url}")
            except requests.RequestException as e:
                logger.debug(f"Request error (attempt {attempt+1}): {str(e)[:80]}")
                time.sleep(5 * (attempt + 1))
        return None

    def _hash(self, company):
        key = f"{company.get('nameAr','')}{company.get('nameEn','')}{company.get('phone1','')}".strip().lower()
        return hashlib.md5(key.encode()).hexdigest()

    def _add(self, company, source='unknown'):
        if len(self.companies) >= self.max_companies:
            return False
        if not company.get('nameAr') and not company.get('nameEn'):
            return False

        h = self._hash(company)
        if h in self.seen_hashes:
            return False

        self.seen_hashes.add(h)
        company['id'] = f"comp_{len(self.companies)+1:06d}"
        company['source'] = source
        company['lastUpdated'] = datetime.now().strftime('%Y-%m-%d')
        company['priority'] = company.get('priority') or self._classify(company)
        self.companies.append(company)
        self.stats[source] = self.stats.get(source, 0) + 1
        return True

    def _classify(self, c):
        fleet = c.get('fleetSize', 0) or 0
        sector = c.get('sector', '')
        high_priority_sectors = ['transport', 'food', 'petroleum', 'construction',
                                  'public_transport', 'rental', 'distribution', 'delivery']
        if fleet >= 100 or (sector in high_priority_sectors and fleet >= 30):
            return 'A'
        elif fleet >= 10 or sector in high_priority_sectors:
            return 'B'
        return 'C'

    def _detect_city(self, text):
        if not text:
            return 'cairo'
        text_lower = text.lower()
        for key, info in AREAS.items():
            if info['ar'] in text or info['en'].lower() in text_lower:
                return key
        return 'cairo'

    def _extract_phones(self, text):
        """Extract Egyptian phone numbers from text."""
        phones = re.findall(
            r'(?:\+?20[\s\-.]?)?(?:0?2[\s\-.]?\d{4}[\s\-.]?\d{4}|'
            r'0?1[0125][\s\-.]?\d{4}[\s\-.]?\d{4}|'
            r'19\d{3}|16\d{3}|15\d{3})',
            text
        )
        clean = []
        for p in phones:
            p = re.sub(r'[\s\-.]', '', p)
            if len(p) >= 6 and p not in clean:
                clean.append(p)
        return clean[:3]

    def _extract_emails(self, text):
        emails = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', text)
        return [e for e in emails if not any(x in e.lower() for x in
                ['example', 'test', 'sentry', 'wix', 'google', 'facebook', 'wordpress'])]

    def _reached_limit(self):
        return len(self.companies) >= self.max_companies

    # ============================================================
    # SOURCE 1: YELLOW PAGES EGYPT
    # ============================================================
    def scrape_yellowpages(self, max_pages_per_cat=20):
        """Scrape yellowpages.com.eg — potentially thousands of listings."""
        logger.info("=" * 60)
        logger.info("SOURCE 1: YELLOW PAGES EGYPT")
        logger.info("=" * 60)
        initial_count = len(self.companies)

        # Build category URLs to scrape
        base = 'https://www.yellowpages.com.eg'
        all_categories = set()
        for sector_key, sector in SECTORS.items():
            for cat in sector.get('yp_cats', []):
                all_categories.add((cat, sector_key))

        for cat_name, sector_key in all_categories:
            if self._reached_limit():
                break

            logger.info(f"  Category: {cat_name} ({SECTORS[sector_key]['ar']})")

            for page in range(1, max_pages_per_cat + 1):
                if self._reached_limit():
                    break

                # Try multiple URL patterns
                urls_to_try = [
                    f"{base}/en/category/{cat_name}?page={page}",
                    f"{base}/ar/category/{cat_name}?page={page}",
                    f"{base}/search/{cat_name}?page={page}",
                ]

                page_results = 0
                for url in urls_to_try:
                    resp = self._request(url)
                    if not resp:
                        continue

                    soup = BeautifulSoup(resp.text, 'lxml')

                    # Try multiple CSS selectors for listings
                    selectors = [
                        '.company-listing', '.listing-item', '.result-item',
                        '.company-card', '.business-card', '.card',
                        '[itemtype*="LocalBusiness"]', '.listing',
                        'article', '.search-result'
                    ]

                    listings = []
                    for sel in selectors:
                        listings = soup.select(sel)
                        if len(listings) > 1:
                            break

                    if not listings:
                        # Try generic approach - find divs with phone numbers
                        for div in soup.find_all('div', class_=True):
                            text = div.get_text()
                            if re.search(r'(?:02|01[0125])\d{8,}', text.replace(' ', '')):
                                listings.append(div)
                        listings = listings[:50]  # Limit

                    for listing in listings:
                        company = self._parse_yp(listing, sector_key)
                        if company:
                            if self._add(company, 'yellowpages'):
                                page_results += 1

                    if page_results > 0:
                        logger.info(f"    Page {page}: +{page_results} companies")
                        break  # Found working URL pattern
                    elif listings:
                        break  # URL works but no new unique results

                if page_results == 0:
                    break  # No more results for this category

        added = len(self.companies) - initial_count
        logger.info(f"  >> Yellow Pages total: +{added} companies")

    def _parse_yp(self, el, sector_key):
        try:
            company = {'sector': sector_key}
            text = el.get_text(separator=' ', strip=True)

            # Name: try specific selectors then fallback
            for sel in ['h2', 'h3', 'h4', '.company-name', '.title', 'a.name', '.listing-title']:
                name_el = el.select_one(sel)
                if name_el:
                    name = name_el.get_text(strip=True)
                    if len(name) > 2 and len(name) < 150:
                        # Detect language
                        if any('\u0600' <= c <= '\u06FF' for c in name):
                            company['nameAr'] = name
                        else:
                            company['nameEn'] = name
                        break

            # Phone
            phone_el = el.select_one('.phone, .tel, [href^="tel:"], .phone-number')
            if phone_el:
                phone = phone_el.get_text(strip=True) or phone_el.get('href', '').replace('tel:', '')
                company['phone1'] = re.sub(r'[\s\-.]', '', phone)
            else:
                phones = self._extract_phones(text)
                if phones:
                    company['phone1'] = phones[0]
                    if len(phones) > 1:
                        company['phone2'] = phones[1]

            # Address & City
            for sel in ['.address', '.location', '.addr', '.area']:
                addr_el = el.select_one(sel)
                if addr_el:
                    company['address'] = addr_el.get_text(strip=True)
                    company['city'] = self._detect_city(company['address'])
                    break
            if not company.get('city'):
                company['city'] = self._detect_city(text)

            # Website
            for a in el.select('a[href]'):
                href = a.get('href', '')
                if href.startswith('http') and 'yellowpages' not in href and 'facebook' not in href:
                    company['website'] = href
                    break

            # Email
            email_el = el.select_one('[href^="mailto:"]')
            if email_el:
                company['email'] = email_el.get('href', '').replace('mailto:', '')
            else:
                emails = self._extract_emails(text)
                if emails:
                    company['email'] = emails[0]

            return company if (company.get('nameAr') or company.get('nameEn')) else None
        except Exception:
            return None

    # ============================================================
    # SOURCE 2: GOOGLE SEARCH
    # ============================================================
    def scrape_google(self, queries_per_sector=3, results_per_query=30):
        """Use Google search to discover companies."""
        logger.info("=" * 60)
        logger.info("SOURCE 2: GOOGLE SEARCH")
        logger.info("=" * 60)
        initial_count = len(self.companies)

        for sector_key, sector in SECTORS.items():
            if self._reached_limit():
                break

            queries_ar = sector.get('google_queries_ar', [])[:queries_per_sector]
            queries_en = sector.get('google_queries_en', [])[:queries_per_sector]

            for area_key, area in list(AREAS.items())[:8]:  # Top 8 areas
                if self._reached_limit():
                    break

                for query_list in [queries_ar, queries_en]:
                    for base_query in query_list:
                        if self._reached_limit():
                            break

                        full_query = f"{base_query} {area['ar']}"

                        for start in range(0, results_per_query, 10):
                            if self._reached_limit():
                                break

                            url = (f"https://www.google.com/search?"
                                   f"q={quote_plus(full_query)}&start={start}&num=10&hl=ar")

                            resp = self._request(url)
                            if not resp:
                                break

                            soup = BeautifulSoup(resp.text, 'lxml')
                            results = soup.select('.g, .tF2Cxc, .MjjYud')

                            if not results:
                                break

                            for result in results:
                                company = self._parse_google_result(result, sector_key, area_key)
                                if company:
                                    self._add(company, 'google')

        added = len(self.companies) - initial_count
        logger.info(f"  >> Google total: +{added} companies")

    def _parse_google_result(self, el, sector_key, area_key):
        try:
            company = {'sector': sector_key, 'city': area_key}

            title_el = el.select_one('h3')
            link_el = el.select_one('a')
            snippet_el = el.select_one('.VwiC3b, .st, .lEBKkf')

            if not title_el:
                return None

            name = title_el.get_text(strip=True)
            # Clean up Google result titles
            name = re.sub(r'\s*[-|–—]\s*(Facebook|LinkedIn|Wuzzuf|فيسبوك|YouTube|تويتر).*$', '', name)
            name = re.sub(r'\s*[-|–—]\s*(الصفحة الرئيسية|Home|About).*$', '', name)

            if len(name) < 3 or len(name) > 100:
                return None

            # Skip non-company results
            skip_domains = ['wikipedia', 'youtube', 'twitter', 'reddit', 'quora', 'pinterest']
            link = link_el.get('href', '') if link_el else ''
            if any(d in link.lower() for d in skip_domains):
                return None

            if any('\u0600' <= c <= '\u06FF' for c in name):
                company['nameAr'] = name
            else:
                company['nameEn'] = name

            if link and link.startswith('http'):
                company['website'] = link

            if snippet_el:
                snippet = snippet_el.get_text(strip=True)
                phones = self._extract_phones(snippet)
                if phones:
                    company['phone1'] = phones[0]
                emails = self._extract_emails(snippet)
                if emails:
                    company['email'] = emails[0]
                company['city'] = self._detect_city(snippet)

            return company
        except Exception:
            return None

    # ============================================================
    # SOURCE 3: EGX LISTED COMPANIES
    # ============================================================
    def scrape_egx(self):
        """Scrape Egyptian Exchange listed companies."""
        logger.info("=" * 60)
        logger.info("SOURCE 3: EGX LISTED COMPANIES")
        logger.info("=" * 60)
        initial_count = len(self.companies)

        url = 'https://www.egx.com.eg/en/ListedStocks.aspx'
        resp = self._request(url)
        if not resp:
            logger.warning("  Could not access EGX website")
            return

        soup = BeautifulSoup(resp.text, 'lxml')
        table = soup.select_one('table.table, #tblListedStock, table')

        if not table:
            logger.warning("  Could not find stock listing table")
            return

        rows = table.select('tr')[1:]  # Skip header
        for row in rows:
            if self._reached_limit():
                break
            cells = row.select('td')
            if len(cells) >= 2:
                name = cells[1].get_text(strip=True) if len(cells) > 1 else cells[0].get_text(strip=True)
                code = cells[0].get_text(strip=True) if len(cells) > 1 else ''

                if name and len(name) > 2:
                    company = {
                        'nameEn': name,
                        'sector': 'manufacturing',
                        'city': 'cairo',
                        'companySize': 'large',
                        'source': 'egx',
                        'notes': f'EGX Code: {code}' if code else 'EGX Listed'
                    }
                    self._add(company, 'egx')

        added = len(self.companies) - initial_count
        logger.info(f"  >> EGX total: +{added} companies")

    # ============================================================
    # SOURCE 4: WUZZUF JOBS (Companies hiring drivers)
    # ============================================================
    def scrape_wuzzuf(self, max_pages=20):
        """Companies posting driver/fleet jobs on Wuzzuf = have vehicle fleets."""
        logger.info("=" * 60)
        logger.info("SOURCE 4: WUZZUF (Companies hiring drivers)")
        logger.info("=" * 60)
        initial_count = len(self.companies)

        keywords = [
            'سائق', 'driver', 'fleet+manager', 'مدير+أسطول',
            'transport+manager', 'logistics', 'سائق+نقل+ثقيل',
            'سائق+تريلا', 'مشرف+حركة', 'warehouse+manager'
        ]

        for keyword in keywords:
            if self._reached_limit():
                break

            for page in range(1, max_pages + 1):
                if self._reached_limit():
                    break

                url = f"https://wuzzuf.net/search/jobs/?q={keyword}&a=hpb&start={page - 1}"
                resp = self._request(url)
                if not resp:
                    break

                soup = BeautifulSoup(resp.text, 'lxml')
                jobs = soup.select('.css-1gatmva, .css-pkv5jc, .job-listing, article')

                if not jobs:
                    break

                for job in jobs:
                    company_el = job.select_one('.css-17s97q8, .css-xkh2o3, .company-name, a[href*="/jobs/companies/"]')
                    if company_el:
                        company_name = company_el.get_text(strip=True)
                        if company_name and len(company_name) > 2:
                            # Detect location
                            location_el = job.select_one('.css-5wys0k, .location, .job-location')
                            city = 'cairo'
                            if location_el:
                                city = self._detect_city(location_el.get_text())

                            company = {
                                'nameEn': company_name,
                                'sector': 'transport',
                                'city': city,
                                'notes': f'Found via job listing: {keyword}'
                            }

                            # Try company link for more info
                            link_el = company_el if company_el.name == 'a' else company_el.find_parent('a')
                            if link_el and link_el.get('href'):
                                company['wuzzufUrl'] = urljoin('https://wuzzuf.net', link_el['href'])

                            self._add(company, 'wuzzuf')

                logger.info(f"  Wuzzuf '{keyword}' page {page}: {len(jobs)} jobs found")

        added = len(self.companies) - initial_count
        logger.info(f"  >> Wuzzuf total: +{added} companies")

    # ============================================================
    # SOURCE 5: BUSINESS DIRECTORIES
    # ============================================================
    def scrape_directories(self, max_pages=10):
        """Scrape various Egyptian business directories."""
        logger.info("=" * 60)
        logger.info("SOURCE 5: BUSINESS DIRECTORIES")
        logger.info("=" * 60)
        initial_count = len(self.companies)

        directories = [
            {
                'name': 'Dalil Misr',
                'base_url': 'https://www.dalil.com/egypt',
                'categories': ['companies', 'factories', 'transport', 'food']
            },
            {
                'name': 'Egypt Industry Directory',
                'base_url': 'https://www.industryegypt.com',
                'categories': ['companies']
            },
            {
                'name': 'Cairo Directory',
                'base_url': 'https://www.cairoegypt.info',
                'categories': ['business']
            },
        ]

        for directory in directories:
            if self._reached_limit():
                break

            logger.info(f"  Trying: {directory['name']}...")

            for cat in directory['categories']:
                if self._reached_limit():
                    break

                for page in range(1, max_pages + 1):
                    url = f"{directory['base_url']}/{cat}?page={page}"
                    resp = self._request(url)
                    if not resp:
                        break

                    soup = BeautifulSoup(resp.text, 'lxml')
                    page_companies = 0

                    # Generic extraction: find anything that looks like a business listing
                    for el in soup.select('article, .listing, .company, .card, .result'):
                        company = self._parse_generic_listing(el)
                        if company and self._add(company, 'directories'):
                            page_companies += 1

                    # Fallback: extract from structured data
                    for script in soup.select('script[type="application/ld+json"]'):
                        try:
                            data = json.loads(script.string)
                            if isinstance(data, dict) and data.get('@type') in ['LocalBusiness', 'Organization']:
                                company = {
                                    'nameEn': data.get('name', ''),
                                    'phone1': data.get('telephone', ''),
                                    'address': data.get('address', {}).get('streetAddress', '') if isinstance(data.get('address'), dict) else str(data.get('address', '')),
                                    'email': data.get('email', ''),
                                    'website': data.get('url', ''),
                                }
                                if company['nameEn']:
                                    company['city'] = self._detect_city(company.get('address', ''))
                                    self._add(company, 'directories')
                                    page_companies += 1
                        except (json.JSONDecodeError, AttributeError):
                            pass

                    if page_companies == 0:
                        break
                    logger.info(f"    Page {page}: +{page_companies}")

        added = len(self.companies) - initial_count
        logger.info(f"  >> Directories total: +{added} companies")

    def _parse_generic_listing(self, el):
        try:
            text = el.get_text(separator=' ', strip=True)
            if len(text) < 10:
                return None

            company = {}

            # Name
            for sel in ['h2', 'h3', 'h4', '.name', '.title', 'a']:
                name_el = el.select_one(sel)
                if name_el:
                    name = name_el.get_text(strip=True)
                    if 3 < len(name) < 150:
                        if any('\u0600' <= c <= '\u06FF' for c in name):
                            company['nameAr'] = name
                        else:
                            company['nameEn'] = name
                        break

            if not company.get('nameAr') and not company.get('nameEn'):
                return None

            phones = self._extract_phones(text)
            if phones:
                company['phone1'] = phones[0]
            emails = self._extract_emails(text)
            if emails:
                company['email'] = emails[0]
            company['city'] = self._detect_city(text)

            return company
        except Exception:
            return None

    # ============================================================
    # SOURCE 6: GOOGLE MAPS SIMULATION (via Google search)
    # ============================================================
    def scrape_google_maps_via_search(self, queries_per_sector=2):
        """Find Google Maps business listings via Google search."""
        logger.info("=" * 60)
        logger.info("SOURCE 6: GOOGLE MAPS (via search)")
        logger.info("=" * 60)
        initial_count = len(self.companies)

        for sector_key, sector in SECTORS.items():
            if self._reached_limit():
                break

            for area_key, area in list(AREAS.items())[:5]:
                if self._reached_limit():
                    break

                queries = sector.get('google_queries_ar', [])[:queries_per_sector]
                for query in queries:
                    if self._reached_limit():
                        break

                    search_query = f"{query} {area['ar']} site:google.com/maps"
                    url = f"https://www.google.com/search?q={quote_plus(search_query)}&num=20"

                    resp = self._request(url)
                    if not resp:
                        continue

                    soup = BeautifulSoup(resp.text, 'lxml')

                    # Look for map pack results and knowledge panels
                    for result in soup.select('.g, .VkpGBb, .cXedhc'):
                        text = result.get_text(separator=' ', strip=True)
                        name_el = result.select_one('h3, .qBF1Pd, .OSrXXb')
                        if name_el:
                            name = name_el.get_text(strip=True)
                            if 3 < len(name) < 100:
                                company = {
                                    'sector': sector_key,
                                    'city': area_key,
                                }
                                if any('\u0600' <= c <= '\u06FF' for c in name):
                                    company['nameAr'] = name
                                else:
                                    company['nameEn'] = name

                                phones = self._extract_phones(text)
                                if phones:
                                    company['phone1'] = phones[0]

                                self._add(company, 'google')

        added = len(self.companies) - initial_count
        logger.info(f"  >> Google Maps total: +{added} companies")

    # ============================================================
    # SOURCE 7: INDUSTRIAL ZONES
    # ============================================================
    def scrape_industrial_zones(self):
        """Search for companies in Egyptian industrial zones."""
        logger.info("=" * 60)
        logger.info("SOURCE 7: INDUSTRIAL ZONES")
        logger.info("=" * 60)
        initial_count = len(self.companies)

        zones = [
            ('المنطقة الصناعية 6 أكتوبر', '6october'),
            ('المنطقة الصناعية العاشر من رمضان', '10thramadan'),
            ('المنطقة الصناعية العبور', 'obour'),
            ('المنطقة الصناعية بدر', 'badr'),
            ('المنطقة الصناعية السادات', 'sadat'),
            ('المنطقة الصناعية أبو رواش', 'giza'),
            ('المنطقة الصناعية حلوان', 'helwan'),
            ('المنطقة الصناعية شبرا الخيمة', 'shubra'),
            ('6th October Industrial Zone companies', '6october'),
            ('10th Ramadan Industrial Zone factories', '10thramadan'),
        ]

        for zone_query, area_key in zones:
            if self._reached_limit():
                break

            # Google search for companies in this zone
            for start in range(0, 40, 10):
                if self._reached_limit():
                    break

                url = f"https://www.google.com/search?q={quote_plus(zone_query)}&start={start}&num=10"
                resp = self._request(url)
                if not resp:
                    break

                soup = BeautifulSoup(resp.text, 'lxml')
                for result in soup.select('.g'):
                    company = self._parse_google_result(result, 'manufacturing', area_key)
                    if company:
                        self._add(company, 'industrial_zones')

        added = len(self.companies) - initial_count
        logger.info(f"  >> Industrial Zones total: +{added} companies")

    # ============================================================
    # SAVE & RESUME
    # ============================================================
    def save_progress(self):
        """Save current progress for resuming later."""
        with open(COMPANIES_CACHE, 'w', encoding='utf-8') as f:
            json.dump(self.companies, f, ensure_ascii=False, indent=2)

        with open(PROGRESS_FILE, 'w', encoding='utf-8') as f:
            json.dump({
                'timestamp': datetime.now().isoformat(),
                'total_companies': len(self.companies),
                'stats': self.stats,
                'hashes': list(self.seen_hashes)
            }, f, ensure_ascii=False, indent=2)

        logger.info(f"Progress saved: {len(self.companies)} companies")

    def load_progress(self):
        """Load previous progress."""
        if not os.path.exists(COMPANIES_CACHE):
            logger.info("No previous progress found")
            return False

        with open(COMPANIES_CACHE, 'r', encoding='utf-8') as f:
            self.companies = json.load(f)

        if os.path.exists(PROGRESS_FILE):
            with open(PROGRESS_FILE, 'r', encoding='utf-8') as f:
                progress = json.load(f)
                self.stats = progress.get('stats', self.stats)
                self.seen_hashes = set(progress.get('hashes', []))

        logger.info(f"Resumed: {len(self.companies)} companies from previous run")
        return True

    # ============================================================
    # EXPORT
    # ============================================================
    def export_excel(self, filename=None):
        """Export to professional Excel spreadsheet."""
        if not self.companies:
            logger.warning("No companies to export")
            return

        if not filename:
            date_str = datetime.now().strftime('%Y%m%d_%H%M')
            filename = os.path.join(OUTPUT_DIR, f'mega_companies_{date_str}.xlsx')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'الشركات'
        ws.sheet_view.rightToLeft = True

        # Styles
        hf = Font(name='Cairo', bold=True, size=11, color='FFFFFF')
        hfill = PatternFill(start_color='4338CA', end_color='4338CA', fill_type='solid')
        ha = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cf = Font(name='Cairo', size=10)
        ca = Alignment(vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        pfills = {
            'A': PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid'),
            'B': PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid'),
            'C': PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid'),
        }

        headers = [
            '#', 'اسم الشركة (عربي)', 'Company Name (EN)', 'القطاع', 'المنطقة',
            'هاتف 1', 'هاتف 2', 'موبايل', 'البريد', 'الموقع',
            'حجم الأسطول', 'جهة الاتصال', 'المسمى', 'الأولوية',
            'حجم الشركة', 'المصدر', 'ملاحظات'
        ]
        widths = [6, 30, 30, 18, 14, 16, 16, 16, 28, 35, 10, 20, 15, 8, 10, 12, 30]

        for i, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.font = hf
            cell.fill = hfill
            cell.alignment = ha
            cell.border = border
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        sector_labels = {k: v['ar'] for k, v in SECTORS.items()}
        area_labels = {k: v['ar'] for k, v in AREAS.items()}

        sorted_companies = sorted(self.companies,
            key=lambda c: ({'A': 0, 'B': 1, 'C': 2}.get(c.get('priority', 'C'), 2),
                           -(c.get('fleetSize', 0) or 0)))

        for row, c in enumerate(sorted_companies, 2):
            data = [
                row - 1,
                c.get('nameAr', ''),
                c.get('nameEn', ''),
                sector_labels.get(c.get('sector', ''), c.get('sector', '')),
                area_labels.get(c.get('city', ''), c.get('city', '')),
                c.get('phone1', ''),
                c.get('phone2', ''),
                c.get('mobile', ''),
                c.get('email', ''),
                c.get('website', ''),
                c.get('fleetSize', ''),
                c.get('contactPerson', ''),
                c.get('contactTitle', ''),
                c.get('priority', 'C'),
                c.get('companySize', ''),
                c.get('source', ''),
                c.get('notes', '')
            ]
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = cf
                cell.alignment = ca
                cell.border = border

            p = c.get('priority', 'C')
            if p in pfills:
                ws.cell(row=row, column=14).fill = pfills[p]

        wb.save(filename)
        logger.info(f"Excel exported: {filename} ({len(self.companies)} companies)")
        return filename

    def export_json(self, filename=None):
        if not filename:
            date_str = datetime.now().strftime('%Y%m%d_%H%M')
            filename = os.path.join(OUTPUT_DIR, f'mega_companies_{date_str}.json')

        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(self.companies, f, ensure_ascii=False, indent=2)
        logger.info(f"JSON exported: {filename}")
        return filename

    def export_csv(self, filename=None):
        if not filename:
            date_str = datetime.now().strftime('%Y%m%d_%H%M')
            filename = os.path.join(OUTPUT_DIR, f'mega_companies_{date_str}.csv')

        with open(filename, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=[
                'nameAr', 'nameEn', 'sector', 'city', 'phone1', 'phone2',
                'mobile', 'email', 'website', 'fleetSize', 'contactPerson',
                'contactTitle', 'priority', 'companySize', 'source', 'notes'
            ])
            writer.writeheader()
            writer.writerows(self.companies)

        logger.info(f"CSV exported: {filename}")
        return filename

    # ============================================================
    # SUMMARY
    # ============================================================
    def print_summary(self):
        print("\n" + "=" * 70)
        print("MEGA SCRAPER - FINAL SUMMARY")
        print("=" * 70)
        print(f"  Total Companies Collected: {len(self.companies):,}")
        print(f"\n  By Source:")
        for source, count in sorted(self.stats.items(), key=lambda x: x[1], reverse=True):
            if count > 0:
                bar = '#' * min(count // 10 + 1, 40)
                print(f"    {source:20s}: {count:6,} {bar}")

        sector_counts = {}
        for c in self.companies:
            s = c.get('sector', 'unknown')
            sector_counts[s] = sector_counts.get(s, 0) + 1

        print(f"\n  By Sector:")
        for sector, count in sorted(sector_counts.items(), key=lambda x: x[1], reverse=True):
            label = SECTORS.get(sector, {}).get('ar', sector)
            print(f"    {label:25s}: {count:6,}")

        with_phone = len([c for c in self.companies if c.get('phone1')])
        with_email = len([c for c in self.companies if c.get('email')])
        with_website = len([c for c in self.companies if c.get('website')])
        total = max(len(self.companies), 1)

        print(f"\n  Data Completeness:")
        print(f"    Phone:   {with_phone:6,} / {total:,} ({100*with_phone//total}%)")
        print(f"    Email:   {with_email:6,} / {total:,} ({100*with_email//total}%)")
        print(f"    Website: {with_website:6,} / {total:,} ({100*with_website//total}%)")
        print("=" * 70)

    # ============================================================
    # RUN
    # ============================================================
    def run(self, sources=None, resume=False):
        if resume:
            self.load_progress()

        # Load curated data first
        from collect_real_data import CURATED_COMPANIES
        logger.info("Loading curated database...")
        for c in CURATED_COMPANIES:
            self._add(dict(c), 'curated')
        logger.info(f"  Curated: {self.stats.get('curated', 0)} companies")

        available_sources = {
            'yellowpages': self.scrape_yellowpages,
            'google': self.scrape_google,
            'egx': self.scrape_egx,
            'wuzzuf': self.scrape_wuzzuf,
            'directories': self.scrape_directories,
            'maps': self.scrape_google_maps_via_search,
            'industrial': self.scrape_industrial_zones,
        }

        if sources:
            run_sources = {k: v for k, v in available_sources.items() if k in sources}
        else:
            run_sources = available_sources

        for name, func in run_sources.items():
            if self._reached_limit():
                logger.info(f"Reached max limit of {self.max_companies} companies. Stopping.")
                break

            try:
                func()
                self.save_progress()  # Save after each source
            except KeyboardInterrupt:
                logger.info("Interrupted by user. Saving progress...")
                self.save_progress()
                break
            except Exception as e:
                logger.error(f"Error in source {name}: {e}")
                self.save_progress()

        self.print_summary()
        self.export_excel()
        self.export_json()
        self.export_csv()

        logger.info("\nDone! Check the 'output' folder.")


def main():
    parser = argparse.ArgumentParser(
        description='Egypt MEGA Scraper - Collect 5,000-30,000 companies',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python mega_scraper.py                              # All sources, up to 30,000 companies
  python mega_scraper.py --max-companies 5000          # Stop at 5,000
  python mega_scraper.py --source yellowpages google   # Only Yellow Pages + Google
  python mega_scraper.py --source wuzzuf               # Only Wuzzuf job listings
  python mega_scraper.py --resume                      # Continue from last run
  python mega_scraper.py --source egx --max-companies 500  # Only EGX listed companies
        """
    )
    parser.add_argument('--source', nargs='+',
                        choices=['yellowpages', 'google', 'egx', 'wuzzuf',
                                 'directories', 'maps', 'industrial'],
                        help='Specific source(s) to scrape')
    parser.add_argument('--max-companies', type=int, default=30000,
                        help='Maximum companies to collect (default: 30000)')
    parser.add_argument('--resume', action='store_true',
                        help='Resume from previous run')
    parser.add_argument('--output', help='Output filename prefix')
    args = parser.parse_args()

    scraper = MegaScraper(max_companies=args.max_companies)
    scraper.run(sources=args.source, resume=args.resume)


if __name__ == '__main__':
    main()
