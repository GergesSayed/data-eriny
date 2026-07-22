"""
Egypt Fleet Companies Scraper
============================
Collects company data from public sources for tire sales targeting.
Focus: Greater Cairo — companies with vehicle fleets.

Usage:
    python egypt_companies_scraper.py --source yellowpages --sector transport
    python egypt_companies_scraper.py --source all --output companies.xlsx
    python egypt_companies_scraper.py --help

Sources:
    1. Yellow Pages Egypt (yellowpages.com.eg)
    2. Google Search (company websites)
    3. Egyptian Exchange (EGX - listed companies)
"""

import os
import sys
import json
import time
import random
import argparse
import hashlib
from datetime import datetime
from urllib.parse import quote_plus, urljoin

try:
    import requests
    from bs4 import BeautifulSoup
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"Missing dependency: {e}")
    print("Install with: pip install -r requirements.txt")
    sys.exit(1)

from config import (
    SECTORS, CITIES, INDUSTRIAL_ZONES, REQUEST_DELAY_SECONDS,
    MAX_RETRIES, REQUEST_TIMEOUT, OUTPUT_DIR, OUTPUT_FILENAME,
    YELLOW_PAGES_BASE_URL, YELLOW_PAGES_CATEGORIES,
    PRIORITY_RULES
)


class CompanyScraper:
    """Main scraper class for collecting Egyptian company data."""

    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
                          '(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept-Language': 'ar,en;q=0.9',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'
        })
        self.companies = []
        self.seen_hashes = set()
        os.makedirs(OUTPUT_DIR, exist_ok=True)

    def _request(self, url, method='GET', **kwargs):
        """Make an HTTP request with retry logic and rate limiting."""
        for attempt in range(MAX_RETRIES):
            try:
                time.sleep(REQUEST_DELAY_SECONDS + random.uniform(0, 1))
                response = self.session.request(method, url, timeout=REQUEST_TIMEOUT, **kwargs)
                response.raise_for_status()
                return response
            except requests.RequestException as e:
                print(f"  ⚠ Request failed (attempt {attempt + 1}/{MAX_RETRIES}): {e}")
                if attempt < MAX_RETRIES - 1:
                    time.sleep(REQUEST_DELAY_SECONDS * (attempt + 2))
        return None

    def _company_hash(self, company):
        """Generate a hash to detect duplicates."""
        key = f"{company.get('nameAr', '')}{company.get('nameEn', '')}{company.get('phone1', '')}".lower()
        return hashlib.md5(key.encode()).hexdigest()

    def _add_company(self, company):
        """Add a company if not duplicate."""
        h = self._company_hash(company)
        if h not in self.seen_hashes:
            self.seen_hashes.add(h)
            company['id'] = f"comp_{len(self.companies) + 1:05d}"
            company['lastUpdated'] = datetime.now().strftime('%Y-%m-%d')
            self.companies.append(company)
            return True
        return False

    def classify_priority(self, company):
        """Auto-classify company priority based on rules."""
        fleet = company.get('fleetSize', 0) or 0
        sector = company.get('sector', '')
        size = company.get('companySize', '')

        for priority, rules in PRIORITY_RULES.items():
            if (fleet >= rules['min_fleet_size'] and
                (not rules.get('sectors') or sector in rules['sectors']) and
                (not rules.get('company_sizes') or size in rules['company_sizes'] or not size)):
                return priority
        return 'B'

    # ---- Source: Yellow Pages Egypt ----
    def scrape_yellow_pages(self, category=None, city=None, max_pages=5):
        """Scrape company listings from Yellow Pages Egypt."""
        categories = [category] if category else YELLOW_PAGES_CATEGORIES
        print(f"\n🟡 Scraping Yellow Pages Egypt...")
        total_added = 0

        for cat in categories:
            print(f"\n  📂 Category: {cat}")
            for page in range(1, max_pages + 1):
                url = f"{YELLOW_PAGES_BASE_URL}/en/category/{cat}"
                if page > 1:
                    url += f"?page={page}"
                if city:
                    url += f"{'&' if '?' in url else '?'}city={city}"

                print(f"    Page {page}... ", end='')
                response = self._request(url)
                if not response:
                    print("❌ Failed")
                    break

                soup = BeautifulSoup(response.text, 'lxml')
                listings = soup.select('.company-listing, .listing-item, .result-item, .card')

                if not listings:
                    print("No more results")
                    break

                page_added = 0
                for listing in listings:
                    company = self._parse_yp_listing(listing, cat)
                    if company and self._add_company(company):
                        page_added += 1

                total_added += page_added
                print(f"✅ Added {page_added} companies")

        print(f"\n  📊 Total from Yellow Pages: {total_added} companies")
        return total_added

    def _parse_yp_listing(self, listing, category):
        """Parse a single Yellow Pages listing."""
        try:
            company = {
                'source': 'yellowpages',
                'sector': self._map_yp_category(category)
            }

            # Name
            name_el = listing.select_one('h2, h3, .company-name, .listing-title, a.title')
            if name_el:
                company['nameEn'] = name_el.get_text(strip=True)
                company['nameAr'] = ''

            # Phone
            phone_el = listing.select_one('.phone, .tel, [href^="tel:"], .phone-number')
            if phone_el:
                phone = phone_el.get_text(strip=True) or phone_el.get('href', '').replace('tel:', '')
                company['phone1'] = phone.strip()

            # Address
            addr_el = listing.select_one('.address, .location, .addr')
            if addr_el:
                address = addr_el.get_text(strip=True)
                company['address'] = address
                company['city'] = self._detect_city(address)

            # Website
            web_el = listing.select_one('a[href*="http"]:not([href*="yellowpages"])')
            if web_el:
                company['website'] = web_el.get('href', '')

            # Email
            email_el = listing.select_one('[href^="mailto:"]')
            if email_el:
                company['email'] = email_el.get('href', '').replace('mailto:', '')

            if not company.get('nameEn') and not company.get('nameAr'):
                return None

            company['priority'] = self.classify_priority(company)
            return company

        except Exception as e:
            print(f"    ⚠ Parse error: {e}")
            return None

    def _map_yp_category(self, category):
        """Map Yellow Pages category to our sector codes."""
        mapping = {
            'transport-companies': 'transport',
            'freight-companies': 'transport',
            'logistics': 'distribution',
            'food-manufacturers': 'food',
            'pharmaceutical-companies': 'pharma',
            'construction-companies': 'construction',
            'car-rental': 'rental',
            'security-companies': 'security',
            'tourism-companies': 'tourism',
            'hospitals': 'healthcare',
            'schools': 'education',
            'factories': 'manufacturing',
        }
        return mapping.get(category, 'manufacturing')

    def _detect_city(self, address):
        """Detect city from address text."""
        address_lower = address.lower()
        for key, city_info in CITIES.items():
            if (city_info['ar'] in address or
                city_info['en'].lower() in address_lower):
                return key
        return 'cairo'  # Default

    # ---- Source: Google Search ----
    def search_google(self, query, num_results=20):
        """Search Google for company information.
        Note: This uses Google search which has rate limits.
        For production use, consider Google Custom Search API.
        """
        print(f"\n🔍 Google Search: '{query}'")
        encoded_query = quote_plus(query)
        url = f"https://www.google.com/search?q={encoded_query}&num={num_results}&hl=ar"

        response = self._request(url)
        if not response:
            print("  ❌ Google search failed")
            return []

        soup = BeautifulSoup(response.text, 'lxml')
        results = []

        for result in soup.select('.g, .tF2Cxc'):
            title_el = result.select_one('h3')
            link_el = result.select_one('a')
            snippet_el = result.select_one('.VwiC3b, .st')

            if title_el and link_el:
                results.append({
                    'title': title_el.get_text(strip=True),
                    'url': link_el.get('href', ''),
                    'snippet': snippet_el.get_text(strip=True) if snippet_el else ''
                })

        print(f"  Found {len(results)} results")
        return results

    # ---- Source: Company Website ----
    def scrape_company_website(self, url):
        """Extract contact information from a company website."""
        print(f"  🌐 Scraping: {url}... ", end='')
        response = self._request(url)
        if not response:
            print("❌")
            return {}

        soup = BeautifulSoup(response.text, 'lxml')
        info = {}

        # Extract phone numbers
        import re
        text = soup.get_text()
        phones = re.findall(r'(?:\+?20|0)[\s-]?(?:1[0-9]|2|3|15)[\s-]?\d{3,4}[\s-]?\d{3,4}', text)
        if phones:
            info['phone1'] = phones[0].strip()
            if len(phones) > 1:
                info['phone2'] = phones[1].strip()

        # Extract emails
        emails = re.findall(r'[\w.+-]+@[\w-]+\.[\w.-]+', text)
        if emails:
            info['email'] = emails[0]

        # Extract from meta tags
        description = soup.find('meta', attrs={'name': 'description'})
        if description:
            info['description'] = description.get('content', '')

        # Extract from contact page link
        contact_link = soup.find('a', text=re.compile(r'contact|اتصل|تواصل', re.I))
        if contact_link and contact_link.get('href'):
            contact_url = urljoin(url, contact_link['href'])
            contact_response = self._request(contact_url)
            if contact_response:
                contact_soup = BeautifulSoup(contact_response.text, 'lxml')
                contact_text = contact_soup.get_text()
                contact_phones = re.findall(r'(?:\+?20|0)[\s-]?(?:1[0-9]|2|3|15)[\s-]?\d{3,4}[\s-]?\d{3,4}', contact_text)
                if contact_phones and not info.get('phone1'):
                    info['phone1'] = contact_phones[0].strip()
                contact_emails = re.findall(r'[\w.+-]+@[\w-]+\.[\w.-]+', contact_text)
                if contact_emails and not info.get('email'):
                    info['email'] = contact_emails[0]

        print("✅")
        return info

    # ---- Export to Excel ----
    def export_to_excel(self, filename=None):
        """Export collected companies to a professionally formatted Excel file."""
        if not self.companies:
            print("\n⚠ No companies to export")
            return

        if not filename:
            date_str = datetime.now().strftime('%Y%m%d_%H%M')
            filename = os.path.join(OUTPUT_DIR, f"{OUTPUT_FILENAME}_{date_str}.xlsx")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'الشركات'
        ws.sheet_view.rightToLeft = True

        # Styling
        header_font = Font(name='Cairo', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_font = Font(name='Cairo', size=10)
        cell_alignment = Alignment(vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )

        # Priority colors
        priority_fills = {
            'A': PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid'),
            'B': PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid'),
            'C': PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid'),
        }

        # Headers
        headers = [
            'الرقم', 'اسم الشركة (عربي)', 'اسم الشركة (إنجليزي)', 'القطاع',
            'المنطقة', 'المحافظة', 'العنوان', 'هاتف 1', 'هاتف 2', 'موبايل',
            'البريد الإلكتروني', 'الموقع', 'حجم الأسطول', 'نوع الأسطول',
            'جهة الاتصال', 'المسمى الوظيفي', 'تليفون المسؤول',
            'حجم الشركة', 'الأولوية', 'المصدر', 'آخر تحديث', 'ملاحظات'
        ]

        column_widths = [6, 25, 25, 15, 12, 12, 30, 15, 15, 15, 25, 30, 10, 12, 18, 15, 15, 10, 8, 12, 12, 25]

        for col_idx, (header, width) in enumerate(zip(headers, column_widths), 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Auto-filter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        # Data rows
        for row_idx, company in enumerate(self.companies, 2):
            sector_info = SECTORS.get(company.get('sector', ''), {})
            city_info = CITIES.get(company.get('city', ''), {})

            row_data = [
                row_idx - 1,
                company.get('nameAr', ''),
                company.get('nameEn', ''),
                sector_info.get('ar', company.get('sector', '')),
                city_info.get('ar', company.get('city', '')),
                company.get('governorate', ''),
                company.get('address', ''),
                company.get('phone1', ''),
                company.get('phone2', ''),
                company.get('mobile', ''),
                company.get('email', ''),
                company.get('website', ''),
                company.get('fleetSize', ''),
                company.get('fleetType', ''),
                company.get('contactPerson', ''),
                company.get('contactTitle', ''),
                company.get('contactPhone', ''),
                company.get('companySize', ''),
                company.get('priority', 'B'),
                company.get('source', ''),
                company.get('lastUpdated', ''),
                company.get('notes', '')
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = thin_border

            # Color priority column
            priority = company.get('priority', 'B')
            priority_cell = ws.cell(row=row_idx, column=19)  # Priority column
            if priority in priority_fills:
                priority_cell.fill = priority_fills[priority]

            # Alternate row colors
            if row_idx % 2 == 0:
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if col_idx != 19:  # Don't override priority color
                        cell.fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')

        # Add Sectors reference sheet
        ws2 = wb.create_sheet('القطاعات')
        ws2.sheet_view.rightToLeft = True
        sector_headers = ['الرمز', 'القطاع (عربي)', 'القطاع (إنجليزي)']
        for col_idx, header in enumerate(sector_headers, 1):
            cell = ws2.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for row_idx, (key, val) in enumerate(SECTORS.items(), 2):
            ws2.cell(row=row_idx, column=1, value=key)
            ws2.cell(row=row_idx, column=2, value=val['ar'])

        # Add Cities reference sheet
        ws3 = wb.create_sheet('المناطق')
        ws3.sheet_view.rightToLeft = True
        city_headers = ['الرمز', 'المنطقة (عربي)', 'المنطقة (إنجليزي)']
        for col_idx, header in enumerate(city_headers, 1):
            cell = ws3.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for row_idx, (key, val) in enumerate(CITIES.items(), 2):
            ws3.cell(row=row_idx, column=1, value=key)
            ws3.cell(row=row_idx, column=2, value=val['ar'])
            ws3.cell(row=row_idx, column=3, value=val['en'])

        wb.save(filename)
        print(f"\n✅ Exported {len(self.companies)} companies to: {filename}")
        return filename

    # ---- Export to JSON (for CRM import) ----
    def export_to_json(self, filename=None):
        """Export to JSON format for direct CRM import."""
        if not filename:
            date_str = datetime.now().strftime('%Y%m%d_%H%M')
            filename = os.path.join(OUTPUT_DIR, f"{OUTPUT_FILENAME}_{date_str}.json")

        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(self.companies, f, ensure_ascii=False, indent=2)

        print(f"✅ Exported {len(self.companies)} companies to: {filename}")
        return filename

    # ---- Run All Sources ----
    def run_all(self, max_pages=3):
        """Run all scraping sources."""
        print("=" * 60)
        print("🚀 Egypt Fleet Companies Scraper")
        print(f"   Target: Greater Cairo — Companies with Vehicle Fleets")
        print(f"   Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        print("=" * 60)

        # 1. Yellow Pages
        self.scrape_yellow_pages(max_pages=max_pages)

        # 2. Summary
        print("\n" + "=" * 60)
        print(f"📊 SUMMARY")
        print(f"   Total Companies: {len(self.companies)}")

        sector_counts = {}
        for c in self.companies:
            s = c.get('sector', 'unknown')
            sector_counts[s] = sector_counts.get(s, 0) + 1

        print(f"\n   By Sector:")
        for sector, count in sorted(sector_counts.items(), key=lambda x: x[1], reverse=True):
            sector_name = SECTORS.get(sector, {}).get('ar', sector)
            print(f"     {sector_name}: {count}")

        priority_counts = {'A': 0, 'B': 0, 'C': 0}
        for c in self.companies:
            p = c.get('priority', 'B')
            priority_counts[p] = priority_counts.get(p, 0) + 1

        print(f"\n   By Priority:")
        print(f"     🔴 A (High): {priority_counts['A']}")
        print(f"     🟡 B (Medium): {priority_counts['B']}")
        print(f"     🟢 C (Low): {priority_counts['C']}")
        print("=" * 60)

        # 3. Export
        self.export_to_excel()
        self.export_to_json()


def main():
    parser = argparse.ArgumentParser(
        description='Egypt Fleet Companies Scraper — جمع بيانات شركات الأساطيل في مصر'
    )
    parser.add_argument('--source', choices=['yellowpages', 'google', 'all'],
                        default='all', help='Data source to scrape')
    parser.add_argument('--sector', choices=list(SECTORS.keys()),
                        help='Specific sector to target')
    parser.add_argument('--city', choices=list(CITIES.keys()),
                        help='Specific city to target')
    parser.add_argument('--max-pages', type=int, default=3,
                        help='Maximum pages per category (default: 3)')
    parser.add_argument('--output', help='Output Excel filename')
    parser.add_argument('--json', action='store_true', help='Also export to JSON')

    args = parser.parse_args()

    scraper = CompanyScraper()

    if args.source == 'yellowpages':
        category = None
        if args.sector:
            # Map sector to YP category
            sector_to_yp = {
                'transport': 'transport-companies',
                'food': 'food-manufacturers',
                'pharma': 'pharmaceutical-companies',
                'construction': 'construction-companies',
                'rental': 'car-rental',
                'security': 'security-companies',
                'tourism': 'tourism-companies',
                'healthcare': 'hospitals',
                'education': 'schools',
                'manufacturing': 'factories',
            }
            category = sector_to_yp.get(args.sector)
        scraper.scrape_yellow_pages(category=category, city=args.city, max_pages=args.max_pages)
    elif args.source == 'google':
        for sector_key, sector_info in SECTORS.items():
            if args.sector and sector_key != args.sector:
                continue
            for keyword in sector_info['keywords'][:2]:
                query = f"{keyword} القاهرة" if args.city is None else f"{keyword} {CITIES[args.city]['ar']}"
                scraper.search_google(query)
    else:
        scraper.run_all(max_pages=args.max_pages)

    if scraper.companies:
        scraper.export_to_excel(args.output)
        if args.json:
            scraper.export_to_json()
    else:
        print("\n⚠ No companies were collected. Try different parameters.")


if __name__ == '__main__':
    main()
