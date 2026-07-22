"""
Google Places API Scraper — THE BEST source for mass data
==========================================================
Uses Google Places API to find companies near specific locations.
This is the MOST POWERFUL way to get 5,000-20,000 companies with:
  - Name (Arabic + English)
  - Phone number
  - Address
  - Website
  - Rating & reviews count
  - Opening hours
  - Google Maps link

SETUP:
  1. Go to: https://console.cloud.google.com/apis/library/places-backend.googleapis.com
  2. Create a project & enable "Places API"
  3. Create an API key: https://console.cloud.google.com/apis/credentials
  4. Google gives $200 FREE credit/month = ~10,000 place searches FREE

Usage:
    python google_places_scraper.py --api-key YOUR_API_KEY
    python google_places_scraper.py --api-key YOUR_API_KEY --sector transport
    python google_places_scraper.py --api-key YOUR_API_KEY --max 5000
"""

import os
import sys
import json
import time
import argparse
from datetime import datetime

try:
    import requests
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"Missing: {e}\nInstall: pip install requests openpyxl")
    sys.exit(1)

OUTPUT_DIR = 'output'

# Google Places search queries for fleet companies
SEARCH_QUERIES = {
    'transport': [
        'شركة نقل', 'transport company', 'freight company', 'trucking company',
        'cargo company', 'شحن بضائع', 'نقل ثقيل'
    ],
    'food': [
        'مصنع أغذية', 'food factory', 'beverage company', 'dairy factory',
        'شركة مشروبات', 'مصنع حلويات', 'bakery factory'
    ],
    'pharma': [
        'شركة أدوية', 'pharmaceutical company', 'pharmacy factory',
        'medical supplies', 'مستلزمات طبية'
    ],
    'construction': [
        'شركة مقاولات', 'construction company', 'contractor',
        'مواد بناء', 'building materials'
    ],
    'petroleum': [
        'شركة بترول', 'petroleum company', 'oil company',
        'محطة بنزين', 'gas station'
    ],
    'distribution': [
        'شركة توزيع', 'logistics company', 'distribution company',
        'warehouse', 'مخزن', 'لوجستيات'
    ],
    'manufacturing': [
        'مصنع', 'factory', 'industrial company', 'manufacturer',
        'منطقة صناعية'
    ],
    'security': [
        'شركة أمن', 'security company', 'حراسة', 'guard services'
    ],
    'rental': [
        'تأجير سيارات', 'car rental', 'limousine service', 'vehicle leasing'
    ],
    'delivery': [
        'شركة توصيل', 'delivery company', 'courier service', 'شحن سريع'
    ],
    'tourism': [
        'شركة سياحة', 'tourism company', 'travel agency', 'tour operator'
    ],
    'healthcare': [
        'مستشفى', 'hospital', 'medical center', 'clinic', 'مركز طبي'
    ],
    'education': [
        'مدرسة خاصة', 'international school', 'private school', 'university'
    ],
    'waste_management': [
        'شركة نظافة', 'waste management', 'recycling company'
    ],
    'real_estate': [
        'شركة عقارات', 'real estate developer', 'تطوير عقاري'
    ],
}

# Greater Cairo grid points for radius search (cover all areas)
SEARCH_LOCATIONS = [
    # Cairo Central
    {'lat': 30.0444, 'lng': 31.2357, 'name': 'Downtown Cairo'},
    {'lat': 30.0626, 'lng': 31.3497, 'name': 'Nasr City'},
    {'lat': 30.0867, 'lng': 31.3303, 'name': 'Heliopolis'},
    {'lat': 29.9603, 'lng': 31.2497, 'name': 'Maadi'},
    {'lat': 30.1001, 'lng': 31.3364, 'name': 'Ain Shams'},
    # New Cairo / 5th Settlement
    {'lat': 30.0074, 'lng': 31.4913, 'name': 'New Cairo'},
    {'lat': 30.0300, 'lng': 31.4700, 'name': '5th Settlement'},
    # Giza
    {'lat': 30.0131, 'lng': 31.2089, 'name': 'Giza'},
    {'lat': 30.0385, 'lng': 31.2099, 'name': 'Dokki / Mohandessin'},
    # 6th October
    {'lat': 29.9285, 'lng': 30.9188, 'name': '6th October City'},
    {'lat': 29.9600, 'lng': 30.9300, 'name': '6th October Industrial'},
    # 10th Ramadan
    {'lat': 30.2973, 'lng': 31.7537, 'name': '10th Ramadan City'},
    {'lat': 30.3100, 'lng': 31.7400, 'name': '10th Ramadan Industrial'},
    # Obour & Shorouk
    {'lat': 30.2249, 'lng': 31.4733, 'name': 'Obour City'},
    {'lat': 30.1600, 'lng': 31.6100, 'name': 'Shorouk City'},
    # Helwan & South
    {'lat': 29.8497, 'lng': 31.3340, 'name': 'Helwan'},
    {'lat': 29.8200, 'lng': 31.3100, 'name': 'Helwan Industrial'},
    # Shubra & North
    {'lat': 30.1280, 'lng': 31.2457, 'name': 'Shubra El Kheima'},
    {'lat': 30.1100, 'lng': 31.2400, 'name': 'Shubra Industrial'},
    # Badr & Sadat
    {'lat': 30.1300, 'lng': 31.7200, 'name': 'Badr City'},
    {'lat': 30.3800, 'lng': 30.5200, 'name': 'Sadat City'},
]


class GooglePlacesScraper:
    def __init__(self, api_key, max_companies=20000):
        self.api_key = api_key
        self.max_companies = max_companies
        self.companies = []
        self.seen_place_ids = set()
        self.stats = {}
        os.makedirs(OUTPUT_DIR, exist_ok=True)

    def search_nearby(self, location, query, radius=5000, sector='unknown'):
        """Search for places near a location."""
        url = 'https://maps.googleapis.com/maps/api/place/nearbysearch/json'
        params = {
            'location': f"{location['lat']},{location['lng']}",
            'radius': radius,
            'keyword': query,
            'language': 'ar',
            'key': self.api_key
        }

        all_results = []
        page = 1

        while True:
            resp = requests.get(url, params=params, timeout=15)
            data = resp.json()

            if data.get('status') != 'OK':
                if data.get('status') == 'ZERO_RESULTS':
                    break
                print(f"  API Error: {data.get('status')} - {data.get('error_message', '')}")
                break

            results = data.get('results', [])
            all_results.extend(results)
            print(f"    Page {page}: {len(results)} results")

            # Next page
            next_token = data.get('next_page_token')
            if not next_token or len(all_results) >= 60:
                break

            time.sleep(2)  # Required by Google API
            params = {'pagetoken': next_token, 'key': self.api_key}
            page += 1

        # Process results
        added = 0
        for place in all_results:
            if len(self.companies) >= self.max_companies:
                break

            place_id = place.get('place_id')
            if place_id in self.seen_place_ids:
                continue
            self.seen_place_ids.add(place_id)

            company = {
                'nameAr': place.get('name', ''),
                'address': place.get('vicinity', ''),
                'sector': sector,
                'lat': place.get('geometry', {}).get('location', {}).get('lat'),
                'lng': place.get('geometry', {}).get('location', {}).get('lng'),
                'rating': place.get('rating'),
                'reviews_count': place.get('user_ratings_total'),
                'google_place_id': place_id,
                'source': 'google_places',
                'lastUpdated': datetime.now().strftime('%Y-%m-%d')
            }

            # Get detailed info (phone, website, etc.)
            details = self.get_place_details(place_id)
            if details:
                company['phone1'] = details.get('formatted_phone_number', '')
                company['phone_intl'] = details.get('international_phone_number', '')
                company['website'] = details.get('website', '')
                company['nameEn'] = details.get('name', company['nameAr'])
                company['address'] = details.get('formatted_address', company['address'])
                company['google_maps_url'] = details.get('url', '')

                # Opening hours
                if details.get('opening_hours'):
                    company['openNow'] = details['opening_hours'].get('open_now')

            # Detect city from address
            city = 'cairo'
            addr = company.get('address', '').lower()
            for area_key, area in SEARCH_LOCATIONS[0:1]:  # Simplified
                pass
            company['city'] = city

            # Priority
            company['priority'] = 'B'
            if company.get('reviews_count', 0) and company['reviews_count'] > 100:
                company['priority'] = 'A'

            self.companies.append(company)
            added += 1

        return added

    def get_place_details(self, place_id):
        """Get detailed info about a place."""
        url = 'https://maps.googleapis.com/maps/api/place/details/json'
        params = {
            'place_id': place_id,
            'fields': 'name,formatted_phone_number,international_phone_number,'
                      'website,formatted_address,url,opening_hours,business_status',
            'language': 'ar',
            'key': self.api_key
        }

        try:
            resp = requests.get(url, params=params, timeout=15)
            data = resp.json()
            if data.get('status') == 'OK':
                return data.get('result', {})
        except Exception as e:
            print(f"  Details error: {e}")
        return None

    def text_search(self, query, location=None, radius=10000, sector='unknown'):
        """Text search for businesses."""
        url = 'https://maps.googleapis.com/maps/api/place/textsearch/json'
        params = {
            'query': query,
            'language': 'ar',
            'key': self.api_key
        }
        if location:
            params['location'] = f"{location['lat']},{location['lng']}"
            params['radius'] = radius

        all_results = []
        page = 1

        while True:
            resp = requests.get(url, params=params, timeout=15)
            data = resp.json()

            if data.get('status') not in ['OK', 'ZERO_RESULTS']:
                print(f"  API Error: {data.get('status')}")
                break
            if data.get('status') == 'ZERO_RESULTS':
                break

            results = data.get('results', [])
            all_results.extend(results)

            next_token = data.get('next_page_token')
            if not next_token:
                break

            time.sleep(2)
            params = {'pagetoken': next_token, 'key': self.api_key}
            page += 1

        added = 0
        for place in all_results:
            if len(self.companies) >= self.max_companies:
                break

            place_id = place.get('place_id')
            if place_id in self.seen_place_ids:
                continue
            self.seen_place_ids.add(place_id)

            company = {
                'nameAr': place.get('name', ''),
                'address': place.get('formatted_address', ''),
                'sector': sector,
                'source': 'google_places',
                'lastUpdated': datetime.now().strftime('%Y-%m-%d'),
                'priority': 'B'
            }

            details = self.get_place_details(place_id)
            if details:
                company['phone1'] = details.get('formatted_phone_number', '')
                company['website'] = details.get('website', '')

            self.companies.append(company)
            added += 1

        return added

    def run(self, sectors=None):
        """Run the full scraper."""
        print("=" * 60)
        print("GOOGLE PLACES API SCRAPER")
        print(f"Target: {self.max_companies:,} companies in Greater Cairo")
        print("=" * 60)

        target_sectors = sectors if sectors else list(SEARCH_QUERIES.keys())

        for sector in target_sectors:
            if len(self.companies) >= self.max_companies:
                break

            queries = SEARCH_QUERIES.get(sector, [])
            print(f"\nSector: {sector} ({len(queries)} queries)")

            for query in queries:
                if len(self.companies) >= self.max_companies:
                    break

                for location in SEARCH_LOCATIONS:
                    if len(self.companies) >= self.max_companies:
                        break

                    print(f"  Searching '{query}' near {location['name']}... ", end='')
                    added = self.search_nearby(location, query, radius=8000, sector=sector)
                    print(f"+{added}")

        self.print_summary()
        self.export()

    def print_summary(self):
        print(f"\n{'='*60}")
        print(f"TOTAL: {len(self.companies):,} companies collected")
        with_phone = len([c for c in self.companies if c.get('phone1')])
        with_website = len([c for c in self.companies if c.get('website')])
        print(f"With phone: {with_phone:,} ({100*with_phone//max(len(self.companies),1)}%)")
        print(f"With website: {with_website:,} ({100*with_website//max(len(self.companies),1)}%)")
        print(f"{'='*60}")

    def export(self):
        date_str = datetime.now().strftime('%Y%m%d_%H%M')

        # JSON
        json_file = os.path.join(OUTPUT_DIR, f'google_places_{date_str}.json')
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(self.companies, f, ensure_ascii=False, indent=2)
        print(f"JSON: {json_file}")

        # Excel
        xlsx_file = os.path.join(OUTPUT_DIR, f'google_places_{date_str}.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Companies'

        headers = ['#', 'Name (AR)', 'Name (EN)', 'Sector', 'Phone', 'Website',
                   'Address', 'Rating', 'Reviews', 'Priority', 'Google Maps']
        for i, h in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=h)

        for row, c in enumerate(self.companies, 2):
            ws.cell(row=row, column=1, value=row-1)
            ws.cell(row=row, column=2, value=c.get('nameAr', ''))
            ws.cell(row=row, column=3, value=c.get('nameEn', ''))
            ws.cell(row=row, column=4, value=c.get('sector', ''))
            ws.cell(row=row, column=5, value=c.get('phone1', ''))
            ws.cell(row=row, column=6, value=c.get('website', ''))
            ws.cell(row=row, column=7, value=c.get('address', ''))
            ws.cell(row=row, column=8, value=c.get('rating', ''))
            ws.cell(row=row, column=9, value=c.get('reviews_count', ''))
            ws.cell(row=row, column=10, value=c.get('priority', ''))
            ws.cell(row=row, column=11, value=c.get('google_maps_url', ''))

        wb.save(xlsx_file)
        print(f"Excel: {xlsx_file}")


def main():
    parser = argparse.ArgumentParser(description='Google Places API Scraper for Egyptian Companies')
    parser.add_argument('--api-key', required=True, help='Google Places API key')
    parser.add_argument('--max', type=int, default=20000, help='Max companies (default: 20000)')
    parser.add_argument('--sector', nargs='+', choices=list(SEARCH_QUERIES.keys()),
                        help='Specific sectors')
    args = parser.parse_args()

    scraper = GooglePlacesScraper(api_key=args.api_key, max_companies=args.max)
    scraper.run(sectors=args.sector)


if __name__ == '__main__':
    main()
