"""
Browser-Based Google Maps Scraper
===================================
يفتح Google Maps في البراوزر ويسحب بيانات الشركات مباشرة.
لأن البراوزر مش بيتحظر زي Python requests.

الاستخدام:
    python browser_scraper.py                        # كل القطاعات
    python browser_scraper.py --sector transport     # قطاع النقل فقط
    python browser_scraper.py --max 500              # أول 500 شركة
"""

import os
import sys
import json
import time
import hashlib
import argparse
from datetime import datetime

try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    SELENIUM_AVAILABLE = True
except ImportError:
    SELENIUM_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    pass

OUTPUT_DIR = 'output'

# استعلامات البحث في Google Maps
MAPS_SEARCHES = {
    'transport': {
        'name': 'نقل وشحن',
        'queries': [
            'شركة نقل بضائع', 'شركة شحن', 'نقل ثقيل', 'freight company',
            'transport company', 'logistics company', 'cargo company',
            'شركة نقل مبرد', 'شركة حاويات'
        ]
    },
    'food': {
        'name': 'أغذية ومشروبات',
        'queries': [
            'مصنع أغذية', 'شركة مشروبات', 'مصنع ألبان', 'food factory',
            'beverage company', 'dairy factory', 'مصنع حلويات'
        ]
    },
    'pharma': {
        'name': 'أدوية',
        'queries': [
            'شركة أدوية', 'مصنع أدوية', 'pharmaceutical company',
            'صيدلية مركزية', 'مخزن أدوية', 'مستلزمات طبية'
        ]
    },
    'construction': {
        'name': 'مقاولات',
        'queries': [
            'شركة مقاولات', 'construction company', 'مواد بناء',
            'مصنع أسمنت', 'حديد وصلب', 'building materials'
        ]
    },
    'manufacturing': {
        'name': 'مصانع',
        'queries': [
            'مصنع', 'factory', 'مصانع 6 أكتوبر', 'مصانع العاشر من رمضان',
            'مصانع العبور', 'industrial company', 'مصنع بلاستيك'
        ]
    },
    'security': {
        'name': 'أمن وحراسة',
        'queries': ['شركة أمن', 'security company', 'حراسة', 'guard services']
    },
    'rental': {
        'name': 'تأجير سيارات',
        'queries': ['تأجير سيارات', 'car rental', 'ليموزين', 'limousine']
    },
    'distribution': {
        'name': 'توزيع ولوجستيات',
        'queries': ['شركة توزيع', 'لوجستيات', 'distribution company', 'مخازن']
    },
    'delivery': {
        'name': 'توصيل',
        'queries': ['شركة توصيل', 'شحن سريع', 'courier', 'delivery company']
    },
}

# المناطق في القاهرة الكبرى
AREAS_TO_SEARCH = [
    'القاهرة', '6 أكتوبر', 'العاشر من رمضان', 'العبور', 'مدينة نصر',
    'التجمع الخامس', 'حلوان', 'الجيزة', 'شبرا الخيمة', 'المعادي',
]


class BrowserScraper:
    def __init__(self, headless=False, max_companies=5000):
        self.max_companies = max_companies
        self.companies = []
        self.seen = set()
        os.makedirs(OUTPUT_DIR, exist_ok=True)

        if not SELENIUM_AVAILABLE:
            print("❌ Selenium مش مثبت!")
            print("   شغّل الأمر ده:")
            print("   pip install selenium")
            print("")
            print("   وحمّل ChromeDriver من:")
            print("   https://chromedriver.chromium.org/downloads")
            sys.exit(1)

        options = Options()
        if headless:
            options.add_argument('--headless=new')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--lang=ar')
        options.add_argument('--window-size=1920,1080')
        # تجنب الكشف
        options.add_argument('--disable-blink-features=AutomationControlled')
        options.add_experimental_option('excludeSwitches', ['enable-automation'])
        options.add_experimental_option('useAutomationExtension', False)

        try:
            self.driver = webdriver.Chrome(options=options)
            self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            self.wait = WebDriverWait(self.driver, 15)
            print("✅ Browser started")
        except Exception as e:
            print(f"❌ Could not start Chrome: {e}")
            print("\nTry: pip install selenium webdriver-manager")
            print("Or download ChromeDriver: https://chromedriver.chromium.org")
            sys.exit(1)

    def search_google_maps(self, query, area=''):
        """بحث في Google Maps وسحب النتائج."""
        full_query = f"{query} {area}".strip()
        search_url = f"https://www.google.com/maps/search/{full_query.replace(' ', '+')}"

        print(f"  🗺️ بحث: {full_query}... ", end='', flush=True)

        try:
            self.driver.get(search_url)
            time.sleep(4)

            # Scroll through results
            results_panel = None
            try:
                # Try to find the results panel
                selectors = [
                    'div[role="feed"]',
                    'div.m6QErb',
                    'div.section-result-content'
                ]
                for sel in selectors:
                    try:
                        results_panel = self.driver.find_element(By.CSS_SELECTOR, sel)
                        if results_panel:
                            break
                    except:
                        continue
            except:
                pass

            if results_panel:
                # Scroll to load more results
                for _ in range(5):
                    self.driver.execute_script(
                        'arguments[0].scrollTop = arguments[0].scrollHeight',
                        results_panel
                    )
                    time.sleep(2)

            # Extract results
            added = 0
            listing_selectors = [
                'div.Nv2PK',  # Google Maps business card
                'a.hfpxzc',  # Google Maps link
                'div[jsaction*="mouseover"]',
            ]

            for sel in listing_selectors:
                try:
                    listings = self.driver.find_elements(By.CSS_SELECTOR, sel)
                    if listings and len(listings) > 2:
                        for listing in listings[:30]:
                            company = self._extract_maps_listing(listing)
                            if company and self._add(company):
                                added += 1
                        break
                except:
                    continue

            # Fallback: parse page source
            if added == 0:
                page_source = self.driver.page_source
                added = self._extract_from_source(page_source, query)

            print(f"+{added}")
            return added

        except Exception as e:
            print(f"❌ {str(e)[:50]}")
            return 0

    def _extract_maps_listing(self, element):
        """استخراج بيانات شركة من عنصر في Google Maps."""
        try:
            text = element.text or ''
            if not text or len(text) < 5:
                # Try aria-label
                text = element.get_attribute('aria-label') or ''

            if not text or len(text) < 5:
                return None

            lines = text.split('\n')
            company = {'source': 'google_maps'}

            # First line is usually the name
            name = lines[0].strip()
            if len(name) < 3 or len(name) > 120:
                return None

            if any('\u0600' <= c <= '\u06FF' for c in name):
                company['nameAr'] = name
            else:
                company['nameEn'] = name

            # Parse remaining lines for info
            import re
            full_text = ' '.join(lines)

            # Phone
            phones = re.findall(
                r'(?:\+?20[\s\-.]?)?(?:0?2[\s\-.]?\d{4}[\s\-.]?\d{4}|'
                r'0?1[0125][\s\-.]?\d{4}[\s\-.]?\d{4}|'
                r'19\d{3}|16\d{3})',
                full_text
            )
            if phones:
                company['phone1'] = re.sub(r'[\s\-.]', '', phones[0])

            # Rating
            rating_match = re.search(r'(\d+[.,]\d+)\s*(?:★|\()', full_text)
            if rating_match:
                company['rating'] = rating_match.group(1)

            # Address (usually one of the later lines)
            for line in lines[1:]:
                line = line.strip()
                if any(area in line for area in ['القاهرة', 'الجيزة', 'أكتوبر', 'العاشر', 'مصر', 'Cairo', 'Giza']):
                    company['address'] = line
                    break

            # Link
            try:
                href = element.get_attribute('href')
                if href:
                    company['google_maps_url'] = href
            except:
                pass

            company['city'] = self._detect_city(full_text)
            company['lastUpdated'] = datetime.now().strftime('%Y-%m-%d')

            return company
        except:
            return None

    def _extract_from_source(self, source, sector_hint=''):
        """استخراج بيانات من HTML المصدر."""
        import re
        added = 0

        # Find business names and phones in the page source
        # Google Maps embeds data in specific patterns
        names = re.findall(r'"([^"]{5,80})"[,\s]*"(?:\d{2,3}[\-\s]?\d{4,8})"', source)

        phone_pattern = r'(?:\+?20[\s\-.]?)?(?:0?2[\s\-.]?\d{4}[\s\-.]?\d{4}|0?1[0125][\s\-.]?\d{4}[\s\-.]?\d{4})'
        all_phones = re.findall(phone_pattern, source)
        all_phones = list(set(re.sub(r'[\s\-.]', '', p) for p in all_phones))

        # Try aria-label pattern
        aria_labels = re.findall(r'aria-label="([^"]{5,100})"', source)
        for label in aria_labels:
            if any(skip in label.lower() for skip in ['google', 'search', 'menu', 'close', 'zoom']):
                continue

            company = {
                'source': 'google_maps',
                'lastUpdated': datetime.now().strftime('%Y-%m-%d'),
                'sector': sector_hint
            }

            if any('\u0600' <= c <= '\u06FF' for c in label):
                company['nameAr'] = label
            else:
                company['nameEn'] = label

            company['city'] = self._detect_city(label)

            if self._add(company):
                added += 1

        return added

    def _detect_city(self, text):
        if not text:
            return 'cairo'
        mappings = {
            '6 أكتوبر': '6october', 'اكتوبر': '6october', 'october': '6october',
            'العاشر': '10thramadan', 'ramadan': '10thramadan',
            'العبور': 'obour', 'الشروق': 'shorouk',
            'التجمع': 'new_cairo', 'مدينة نصر': 'nasr_city',
            'المعادي': 'maadi', 'حلوان': 'helwan',
            'الجيزة': 'giza', 'شبرا': 'shubra',
        }
        text_lower = text.lower()
        for keyword, area in mappings.items():
            if keyword in text or keyword.lower() in text_lower:
                return area
        return 'cairo'

    def _add(self, company):
        if len(self.companies) >= self.max_companies:
            return False
        key = f"{company.get('nameAr','')}{company.get('nameEn','')}"
        if not key.strip():
            return False
        h = hashlib.md5(key.lower().encode()).hexdigest()
        if h in self.seen:
            return False
        self.seen.add(h)
        company['id'] = f"b{len(self.companies)+1:06d}"
        self.companies.append(company)
        return True

    def run(self, sectors=None):
        print("=" * 60)
        print("🗺️ Browser Google Maps Scraper")
        print(f"   الهدف: {self.max_companies} شركة")
        print("=" * 60)

        target_sectors = sectors or list(MAPS_SEARCHES.keys())

        try:
            for sector_key in target_sectors:
                if len(self.companies) >= self.max_companies:
                    break

                sector = MAPS_SEARCHES.get(sector_key)
                if not sector:
                    continue

                print(f"\n{'='*50}")
                print(f"📂 {sector['name']}")
                print(f"{'='*50}")

                for query in sector['queries']:
                    if len(self.companies) >= self.max_companies:
                        break

                    for area in AREAS_TO_SEARCH[:5]:
                        if len(self.companies) >= self.max_companies:
                            break
                        self.search_google_maps(query, area)

                # Save progress after each sector
                self._save()

        except KeyboardInterrupt:
            print("\n⚠️ توقف — حفظ البيانات...")

        finally:
            self.driver.quit()

        self._save()
        self._print_summary()

    def _save(self):
        date_str = datetime.now().strftime('%Y%m%d_%H%M')

        # JSON
        json_file = os.path.join(OUTPUT_DIR, f'browser_scrape_{date_str}.json')
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(self.companies, f, ensure_ascii=False, indent=2)

        # CRM Import
        crm_file = os.path.join(OUTPUT_DIR, 'crm_import_ready.json')
        # Merge with existing
        existing = []
        if os.path.exists(crm_file):
            with open(crm_file, 'r', encoding='utf-8') as f:
                existing = json.load(f)

        merged = existing + [c for c in self.companies if c['id'] not in {e.get('id') for e in existing}]
        with open(crm_file, 'w', encoding='utf-8') as f:
            json.dump(merged, f, ensure_ascii=False, indent=2)

        # Excel
        try:
            xlsx_file = os.path.join(OUTPUT_DIR, f'browser_scrape_{date_str}.xlsx')
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'الشركات'
            headers = ['#', 'اسم الشركة (عربي)', 'Company Name', 'القطاع',
                       'المنطقة', 'هاتف', 'العنوان', 'التقييم', 'المصدر']
            for i, h in enumerate(headers, 1):
                ws.cell(row=1, column=i, value=h)
                ws.cell(row=1, column=i).font = Font(bold=True)
            for row, c in enumerate(self.companies, 2):
                ws.cell(row=row, column=1, value=row-1)
                ws.cell(row=row, column=2, value=c.get('nameAr', ''))
                ws.cell(row=row, column=3, value=c.get('nameEn', ''))
                ws.cell(row=row, column=4, value=c.get('sector', ''))
                ws.cell(row=row, column=5, value=c.get('city', ''))
                ws.cell(row=row, column=6, value=c.get('phone1', ''))
                ws.cell(row=row, column=7, value=c.get('address', ''))
                ws.cell(row=row, column=8, value=c.get('rating', ''))
                ws.cell(row=row, column=9, value=c.get('source', ''))
            wb.save(xlsx_file)
        except:
            pass

        print(f"\n💾 محفوظ: {len(self.companies)} شركة")

    def _print_summary(self):
        print(f"\n{'='*60}")
        print(f"📊 النتائج النهائية: {len(self.companies)} شركة")
        sectors = {}
        for c in self.companies:
            s = c.get('sector', '?')
            sectors[s] = sectors.get(s, 0) + 1
        for s, count in sorted(sectors.items(), key=lambda x: x[1], reverse=True):
            name = MAPS_SEARCHES.get(s, {}).get('name', s)
            print(f"  {name}: {count}")
        with_phone = len([c for c in self.companies if c.get('phone1')])
        print(f"\n  مع أرقام: {with_phone}/{len(self.companies)} ({100*with_phone//max(len(self.companies),1)}%)")
        print(f"{'='*60}")


def main():
    parser = argparse.ArgumentParser(description='Browser Google Maps Scraper')
    parser.add_argument('--sector', nargs='+', choices=list(MAPS_SEARCHES.keys()))
    parser.add_argument('--max', type=int, default=5000)
    parser.add_argument('--headless', action='store_true', help='Run without showing browser')
    args = parser.parse_args()

    scraper = BrowserScraper(headless=args.headless, max_companies=args.max)
    scraper.run(sectors=args.sector)


if __name__ == '__main__':
    main()
