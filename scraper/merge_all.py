"""Merge all scraped data into one unified file."""
import os
import json
import hashlib
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

OUTPUT_DIR = 'output'

def merge_all():
    all_companies = []
    seen = set()
    sources = {}

    # Load all JSON files
    for filename in os.listdir(OUTPUT_DIR):
        if not filename.endswith('.json') or filename.startswith('_'):
            continue
        filepath = os.path.join(OUTPUT_DIR, filename)
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
            if isinstance(data, list):
                source_name = filename.replace('.json', '')
                count = 0
                for company in data:
                    key = f"{company.get('nameAr','')}{company.get('nameEn','')}{company.get('phone1','')}".lower().strip()
                    h = hashlib.md5(key.encode()).hexdigest()
                    if h not in seen and (company.get('nameAr') or company.get('nameEn')):
                        seen.add(h)
                        company['id'] = f"m{len(all_companies)+1:06d}"
                        all_companies.append(company)
                        count += 1
                sources[source_name] = count
                print(f"  {filename}: {count} unique companies")
        except Exception as e:
            print(f"  {filename}: Error - {e}")

    print(f"\n{'='*60}")
    print(f"  TOTAL UNIQUE COMPANIES: {len(all_companies)}")
    print(f"{'='*60}")

    # Stats
    with_phone = len([c for c in all_companies if c.get('phone1')])
    with_email = len([c for c in all_companies if c.get('email')])
    with_website = len([c for c in all_companies if c.get('website')])
    total = max(len(all_companies), 1)
    print(f"  With phone:   {with_phone}/{total} ({100*with_phone//total}%)")
    print(f"  With email:   {with_email}/{total} ({100*with_email//total}%)")
    print(f"  With website: {with_website}/{total} ({100*with_website//total}%)")

    # Save merged JSON
    date_str = datetime.now().strftime('%Y%m%d_%H%M')
    merged_json = os.path.join(OUTPUT_DIR, f'ALL_COMPANIES_{date_str}.json')
    with open(merged_json, 'w', encoding='utf-8') as f:
        json.dump(all_companies, f, ensure_ascii=False, indent=2)
    print(f"\n  JSON: {merged_json}")

    # Also save CRM import
    crm_file = os.path.join(OUTPUT_DIR, 'crm_import_ready.json')
    with open(crm_file, 'w', encoding='utf-8') as f:
        json.dump(all_companies, f, ensure_ascii=False, indent=2)
    print(f"  CRM:  {crm_file}")

    # Save merged Excel
    if openpyxl:
        merged_xlsx = os.path.join(OUTPUT_DIR, f'ALL_COMPANIES_{date_str}.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'كل الشركات'
        ws.sheet_view.rightToLeft = True

        hf = Font(name='Arial', bold=True, size=11, color='FFFFFF')
        hfill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
        cf = Font(name='Arial', size=10)
        border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB')
        )
        alt_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')

        headers = ['#', 'اسم الشركة (عربي)', 'Company Name (EN)', 'القطاع',
                   'المنطقة', 'هاتف 1', 'هاتف 2', 'البريد', 'الموقع',
                   'العنوان', 'التقييم', 'المصدر']
        widths = [6, 40, 35, 15, 14, 16, 16, 28, 40, 40, 8, 15]

        for i, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.font = hf
            cell.fill = hfill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        for row, c in enumerate(all_companies, 2):
            data = [
                row-1,
                c.get('nameAr', ''),
                c.get('nameEn', ''),
                c.get('sector', ''),
                c.get('city', ''),
                c.get('phone1', ''),
                c.get('phone2', ''),
                c.get('email', ''),
                c.get('website', ''),
                c.get('address', ''),
                c.get('rating', ''),
                c.get('source', ''),
            ]
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = cf
                cell.border = border
                if row % 2 == 0:
                    cell.fill = alt_fill

        wb.save(merged_xlsx)
        print(f"  Excel: {merged_xlsx}")

    print(f"\n{'='*60}")
    print(f"  Done! {len(all_companies)} companies merged.")
    print(f"{'='*60}")


if __name__ == '__main__':
    merge_all()
